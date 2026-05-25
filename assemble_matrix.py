#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Phase 2: 数据组装脚本
功能：从本地 SQLite 缓存库读取禅道数据，组装成标准需求跟踪矩阵格式（38列）
输出格式严格对齐：1研发中心质控部基础数据\需求表\2026-1-V1版本-01-需求跟踪矩阵.xlsx

修复项（对比模板）：
1. 月度 → 月份第1天的Excel序列号（非project.end）
2. 周数 → 基于project.end的周一边界周数
3. 需求状态 → "原始"（已纳入版本的需求，非story.status）
4. 需求提交时间 → story首次reviewed/linked2project/openedDate
5. 被指派时间 → action表首次assigned / story.openedDate
6. 提测时间 → task有数据的取max(openedDate)，无task取None
7. 新增列 → 反馈创建时间(AK)、被指派时间2(AL) 共38列
8. 对应研发/测试 → 取task assignedTo（优先type=devel/test）
9. 版本范围 → 仅指定project的故事（非全量）
"""

import sys
import os
import sqlite3
import re
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import warnings
warnings.filterwarnings('ignore')
sys.path.append(os.path.dirname(__file__))

from config import LOCAL_CACHE_DB, OUTPUT_DIR, TARGET_PROJECT_IDS

# === 48列表头（原38列 + 10个新效能指标列）===
COLUMN_HEADERS = [
    '月度', '周数', '版本名称', '类型', '项目ID',
    '一级部门', '二级部门', '所属业务模块', '需求类型', '需求状态',
    '产品经理', '需求ID', 'bugID', '优先级', '项目需求评级',
    '是否对接第三方', '是否为个性化', '是否为争议', '需求名称', '需求提交时间',
    '对应研发', '对应测试', '提测时间', '验收是否通过', '未通过分类', '未通过原因',
    '是否延期提测', '备注', '被指派时间', '响应时长', '消耗工时',
    '发版完成时间', '机构名称-zentao_new-zt_feedback.organization', '大区', '分区', '反馈优先级',
    '反馈创建时间', '被指派时间',  # 新增2列（38列）
    # === 新增效能指标字段 ===
    '需求估算工时',               # AM 39 - zt_story.estimate
    '需求评审通过时间',           # AN 40 - zt_story.reviewedDate
    '需求开发开始时间',           # AO 41 - zt_task.min(realStarted)
    '缺陷激活次数',               # AP 42 - zt_bug.activatedCount (关联到story的bug)
    '关联Bug数',                 # AQ 43 - zt_bug count per story
    '关联Bug激活总数',            # AR 44 - zt_bug activatedCount sum per story
    '产品ID',                    # AS 45 - story.product
    '产品名称',                   # AT 46 - zt_product.name
    '项目团队人数',               # AU 47 - zt_team count per project
    '项目团队角色分布',            # AV 48 - zt_team roles per project
]

# === 用户 → 部门映射 ===
DEPT_ID_MAP = {
    197: ('研发部', '门诊服务'),
    201: ('研发部', '其他'),
    196: ('产品与数据中心', '产品中心'),
    199: ('产品与数据中心', '产品中心'),
    136: ('产品与数据中心', '数据中心'),
    137: ('产品与数据中心', '运维部'),
    195: ('产品与数据中心', '运维部'),
    198: ('产品与数据中心', '运维部'),
    179: ('产品与数据中心', '数据中心'),
    165: ('产品与数据中心', '数据中心'),
    164: ('产品与数据中心', '数据中心'),
    161: ('产品与数据中心', '数据中心'),
    162: ('产品与数据中心', '数据中心'),
    65: ('业务部', '咨询'),
    63: ('业务部', '患者管理'),
    64: ('业务部', '我的'),
    67: ('业务部', '我的'),
    69: ('业务部', '预约'),
    70: ('业务部', '转诊'),
    27: ('业务部', '预约'),
    28: ('业务部', '转诊'),
    29: ('业务部', '会诊'),
    43: ('业务部', '转诊'),
    44: ('业务部', '会诊'),
    78: ('业务部', '我的'),
    79: ('业务部', '我的'),
    93: ('业务部', '医技检查'),
    160: ('业务部', '其他'),
    156: ('业务部', '远程门诊'),
    159: ('业务部', '其他'),
    166: ('业务部', '其他'),
    169: ('业务部', '其他'),
    175: ('业务部', '其他'),
    176: ('业务部', '其他'),
    177: ('业务部', '其他'),
    178: ('业务部', '其他'),
    180: ('业务部', '其他'),
    181: ('业务部', '其他'),
    182: ('业务部', '其他'),
    183: ('业务部', '其他'),
    184: ('业务部', '其他'),
    185: ('业务部', '其他'),
    191: ('业务部', '咨询'),
    192: ('业务部', '取单'),
    193: ('业务部', '通用'),
}

USER_DEPT_ID = {}


def load_user_dept_map(user_list):
    global USER_DEPT_ID
    for u in user_list:
        acct = u.get('account', '')
        dept_id = u.get('dept', 0)
        if acct:
            try:
                USER_DEPT_ID[acct] = int(dept_id)
            except (ValueError, TypeError):
                USER_DEPT_ID[acct] = 0


def get_user_dept(account):
    if not account or account in ('', 'closed', '/'):
        return ('/', '/')
    dept_id = USER_DEPT_ID.get(str(account), 0)
    if dept_id and dept_id in DEPT_ID_MAP:
        return DEPT_ID_MAP[dept_id]
    return ('/', '/')


def get_realname(account, user_dict):
    """获取用户真实姓名"""
    if not account or account in ('', 'closed', '/'):
        return '/'
    if account in user_dict:
        return user_dict[account]
    return str(account)


def excel_date_serial(dt):
    """日期转Excel序列号（1900日期系统）"""
    if dt is None:
        return None
    if isinstance(dt, str):
        try:
            dt = datetime.strptime(dt, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            try:
                dt = datetime.strptime(dt, '%Y-%m-%d')
            except ValueError:
                return None
    epoch = datetime(1899, 12, 30)
    delta = dt - epoch
    return int(delta.total_seconds() / 86400)


def first_day_of_month(dt):
    """获取月份第1天"""
    if dt is None:
        return None
    return dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)


def get_week_of_month(proj_name):
    """周数由版本迭代决定：V1→第三周，V2→第五周"""
    if not proj_name:
        return ''
    name = str(proj_name)
    if 'V2' in name or 'V2' in name.replace('【常规版】', ''):
        return '第五周'
    else:
        return '第三周'


def parse_date(dt):
    """解析日期字符串为datetime"""
    if not dt or dt in ('', '0000-00-00 00:00:00', '0000-00-00'):
        return None
    if isinstance(dt, datetime):
        return dt
    if isinstance(dt, str):
        try:
            return datetime.strptime(dt, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            try:
                return datetime.strptime(dt, '%Y-%m-%d')
            except ValueError:
                return None
    return None


def format_version_name(proj_name, proj_end):
    """格式化版本名称：去除【常规版】前缀，日期用两位数字"""
    if not proj_name:
        return ''
    name = str(proj_name)
    name = name.replace('【常规版】', '')
    match = re.search(r'（(\d+)日）$', name)
    if match:
        day = int(match.group(1))
        if proj_end:
            try:
                end_dt = datetime.strptime(proj_end, '%Y-%m-%d')
                month_day = f'{end_dt.month:02d}{end_dt.day:02d}'
            except ValueError:
                month_day = f'{day:04d}'
        else:
            month_day = f'{day:04d}'
        name = re.sub(r'（\d+日）$', f'（{month_day}）', name)
    return name


def contains_third_party(title):
    if not title:
        return '否'
    keywords = ['第三方', '对接', '外部', '接口', '联调', 'his', 'HIS', '医保', 'api', 'API']
    for kw in keywords:
        if kw.lower() in str(title).lower():
            return '是'
    return '否'


def contains_custom(title):
    if not title:
        return '否'
    keywords = ['定制', '个性化', '专属', '私有']
    for kw in keywords:
        if kw in str(title):
            return '是'
    return '否'


def load_data():
    conn = sqlite3.connect(LOCAL_CACHE_DB)
    data = {}
    for name in ['project', 'projectstory', 'story', 'task', 'bug',
                 'user', 'module', 'action', 'feedback', 'team', 'product']:
        data[name] = conn.execute(f'SELECT * FROM zt_{name}').fetchall()
        # 获取列名
        cursor = conn.execute(f'SELECT * FROM zt_{name} LIMIT 1')
        data[f'{name}_cols'] = [desc[0] for desc in cursor.description]
    conn.close()
    return data


def rows_to_dict(rows, cols):
    """将sqlite3 rows转为dict列表"""
    return [dict(zip(cols, row)) for row in rows]


def assemble_matrix(tables, target_project_ids=None):
    print("📦 开始组装需求跟踪矩阵...")

    project_list = rows_to_dict(tables['project'], tables['project_cols'])
    projectstory_list = rows_to_dict(tables['projectstory'], tables['projectstory_cols'])
    story_list = rows_to_dict(tables['story'], tables['story_cols'])
    task_list = rows_to_dict(tables['task'], tables['task_cols'])
    bug_list = rows_to_dict(tables['bug'], tables['bug_cols'])
    action_list = rows_to_dict(tables['action'], tables['action_cols'])
    feedback_list = rows_to_dict(tables['feedback'], tables['feedback_cols'])
    module_list = rows_to_dict(tables['module'], tables['module_cols'])
    user_list = rows_to_dict(tables['user'], tables['user_cols'])
    team_list = rows_to_dict(tables['team'], tables['team_cols'])
    product_list = rows_to_dict(tables['product'], tables['product_cols'])

    # 索引化
    project_dict = {p['id']: p for p in project_list}
    module_dict = {m['id']: m['name'] for m in module_list}
    feedback_dict = {f['id']: f for f in feedback_list}
    user_dict = {u['account']: u.get('realname', u['account']) for u in user_list}
    product_dict = {p['id']: p for p in product_list}

    load_user_dept_map(user_list)

    # 项目团队统计
    project_team_count = defaultdict(int)
    project_team_roles = defaultdict(lambda: defaultdict(int))
    for t in team_list:
        root = t.get('root', 0)
        role = t.get('role', '')
        project_team_count[root] += 1
        if role:
            project_team_roles[root][role] += 1

    # story → 关联bug统计（激活次数、bug数、激活总数）
    story_bug_activated = {}
    story_bug_count = defaultdict(int)
    story_bug_activated_sum = defaultdict(int)
    valid_story_ids = {s['id'] for s in story_list}
    for b in bug_list:
        sid = b.get('story', 0)
        if sid and sid > 0 and sid in valid_story_ids:
            story_bug_count[sid] += 1
            act = b.get('activatedCount', 0)
            if act:
                try:
                    act_int = int(act)
                    story_bug_activated_sum[sid] += act_int
                    if act_int > 0:
                        story_bug_activated[sid] = act_int
                except (ValueError, TypeError):
                    pass

    # 目标项目过滤
    if target_project_ids:
        proj_ids = set(target_project_ids)
    else:
        proj_ids = set(project_dict.keys())

    # story → 项目映射
    story_to_project = defaultdict(list)
    for ps in projectstory_list:
        story_to_project[ps['story']].append(ps['project'])

    # story → 开发者（通过task assignedTo，优先type=devel）
    story_dev_map = {}
    for t in task_list:
        sid = t.get('story', 0)
        if not sid or sid <= 0:
            continue
        assigned = t.get('assignedTo', '')
        if not assigned or assigned in ('', 'closed'):
            continue
        if sid not in story_dev_map:
            story_dev_map[sid] = assigned
        elif t.get('type') == 'devel':
            story_dev_map[sid] = assigned

    # story → 测试人员（task type='test' 或 名称含'测试'）
    story_qa_map = {}
    for t in task_list:
        sid = t.get('story', 0)
        if not sid or sid <= 0:
            continue
        assigned = t.get('assignedTo', '')
        if not assigned or assigned in ('', 'closed'):
            continue
        t_type = t.get('type', '')
        t_name = t.get('name', '')
        if t_type == 'test' or '测试' in str(t_name):
            if sid not in story_qa_map:
                story_qa_map[sid] = assigned

    # story → bugID（通过 action frombug）
    story_bug_map = {}
    for a in action_list:
        if a.get('action') == 'frombug':
            sid = a.get('objectID', 0)
            if sid and sid not in story_bug_map:
                try:
                    story_bug_map[sid] = int(a.get('extra', 0))
                except (ValueError, TypeError):
                    pass

    # story → 总消耗工时
    story_consumed = defaultdict(float)
    for t in task_list:
        sid = t.get('story', 0)
        consumed = t.get('consumed', 0)
        if sid and consumed:
            try:
                story_consumed[sid] += float(consumed)
            except (ValueError, TypeError):
                pass

    # story → 第一次被指派时间（action表）
    story_first_assigned = {}
    for a in action_list:
        if a.get('action') == 'assigned':
            sid = a.get('objectID', 0)
            date_val = a.get('date', '')
            if sid and date_val and sid not in story_first_assigned:
                story_first_assigned[sid] = date_val

    # story → 第一次reviewed时间
    story_first_reviewed = {}
    for a in action_list:
        if a.get('action') == 'reviewed':
            sid = a.get('objectID', 0)
            date_val = a.get('date', '')
            if sid and date_val and sid not in story_first_reviewed:
                story_first_reviewed[sid] = date_val

    # story → 第一次linked2project时间
    story_first_linked = {}
    for a in action_list:
        if a.get('action') == 'linked2project':
            sid = a.get('objectID', 0)
            date_val = a.get('date', '')
            if sid and date_val and sid not in story_first_linked:
                story_first_linked[sid] = date_val

    # story → 是否被unlinkedfromproject（记录被删除的story）
    story_unlinked = {}
    for a in action_list:
        if a.get('action') == 'unlinkedfromproject':
            sid = a.get('objectID', 0)
            if sid and sid not in story_unlinked:
                story_unlinked[sid] = a.get('extra', '')

    # story → 第一次reviewmodify时间（修改后重新评审通过）
    story_reviewmodify = {}
    for a in action_list:
        if a.get('action') == 'reviewmodify':
            sid = a.get('objectID', 0)
            date_val = a.get('date', '')
            if sid and date_val and sid not in story_reviewmodify:
                story_reviewmodify[sid] = date_val

    # story → 第一次frombug的bugID（extra字段）
    for a in action_list:
        if a.get('action') == 'frombug':
            sid = a.get('objectID', 0)
            if sid and sid not in story_bug_map:
                try:
                    story_bug_map[sid] = int(a.get('extra', 0))
                except (ValueError, TypeError):
                    pass

    # story → 是否有makemodifypostpone（延期标记）
    story_postponed = {}
    for a in action_list:
        if a.get('action') == 'makemodifypostpone':
            sid = a.get('objectID', 0)
            if sid:
                story_postponed[sid] = True

    # story → 关联task的最大openedDate（提测时间候选）
    story_max_task_opened = {}
    # story → 关联task的最小realStarted（开发开始时间）
    story_min_real_started = {}
    for t in task_list:
        sid = t.get('story', 0)
        opened = t.get('openedDate', '')
        if sid and opened and opened not in ('', '0000-00-00 00:00:00'):
            if sid not in story_max_task_opened or opened > story_max_task_opened[sid]:
                story_max_task_opened[sid] = opened
        real_started = t.get('realStarted', '')
        if sid and real_started and real_started not in ('', '0000-00-00 00:00:00'):
            if sid not in story_min_real_started or real_started < story_min_real_started[sid]:
                story_min_real_started[sid] = real_started

    # feedback → 创建时间（feedback.openedDate）
    fb_created = {f['id']: f.get('openedDate', '') for f in feedback_list if f.get('id')}

    # === 组装每一行 ===
    rows = []
    processed_story_ids = set()

    for story in story_list:
        story_id = story['id']

        # 过滤已删除的story
        if story.get('deleted', 0) in (1, '1', True):
            continue

        # 找到关联的 target 项目
        linked_projects = story_to_project.get(story_id, [])
        target_proj = None
        for pid in linked_projects:
            if pid in proj_ids and pid in project_dict:
                target_proj = project_dict[pid]
                break
        if not target_proj:
            continue

        if story_id in processed_story_ids:
            continue
        processed_story_ids.add(story_id)

        project_id = target_proj['id']

        # === 解析日期 ===
        proj_end_dt = parse_date(target_proj.get('end', ''))
        proj_begin_dt = parse_date(target_proj.get('begin', ''))
        faban_dt = proj_end_dt

        # A: 月度 → Excel日期序列号 + 数字格式 "yyyy/m/d" 显示为 "2026/1/1"
        month_first = first_day_of_month(proj_end_dt)
        month_serial = excel_date_serial(month_first)

        # B: 周数（基于project.end的周一边界）
        week_name = get_week_of_month(target_proj.get('name', ''))

        # C: 版本名称
        version_name = format_version_name(target_proj.get('name', ''), target_proj.get('end', ''))

        # D: 类型
        type_name = '常规版' if target_proj.get('type') == 'sprint' else str(target_proj.get('type', ''))

        # E: 项目ID
        proj_id = int(project_id)

        # F/G: 一级/二级部门（基于开发者）
        dev_account = story_dev_map.get(story_id, story.get('assignedTo', ''))
        dept1, dept2 = get_user_dept(dev_account)

        # H: 所属业务模块
        mod_id = story.get('module', 0)
        mod_name = module_dict.get(mod_id, '/') if mod_id and int(mod_id) > 0 else '/'

        # I: 需求类型
        req_type = '项目需求'

        # J: 需求状态 → 原始/新增/删除/延期
        if story_id in story_unlinked:
            still_linked = project_id in linked_projects
            if not still_linked:
                status = '删除'
            else:
                # 已重新链接
                first_date = story_first_reviewed.get(story_id) or story_first_linked.get(story_id, '')
                if first_date:
                    try:
                        dt = datetime.strptime(str(first_date)[:10], '%Y-%m-%d')
                        if dt <= datetime(2025, 12, 31):
                            status = '原始'
                        else:
                            status = '新增'
                    except ValueError:
                        status = '原始'
                else:
                    status = '原始'
        else:
            # 从未被unlinked，检查是否有makemodifypostpone（延期）
            if story_id in story_postponed:
                status = '延期'
            else:
                # 首次reviewed/linked日期：2025-12-31前为原始，之后为新增
                first_date = story_first_reviewed.get(story_id) or story_first_linked.get(story_id, '')
                if first_date:
                    try:
                        dt = datetime.strptime(str(first_date)[:10], '%Y-%m-%d')
                        if dt <= datetime(2025, 12, 31):
                            status = '原始'
                        else:
                            status = '新增'
                    except ValueError:
                        status = '原始'
                else:
                    status = '原始'

        # K: 产品经理
        pm_account = story.get('reviewedBy', '')
        pm_name = get_realname(pm_account, user_dict)

        # L: 需求ID
        req_id = int(story_id)

        # M: bugID
        bug_id = story_bug_map.get(story_id, None)

        # N: 优先级
        pri_map = {0: 'P0', 1: 'P0', 2: 'P1', 3: 'P2', 4: 'P3'}
        pri_val = 0
        try:
            pri_val = int(story.get('pri', 0) or 0)
        except (ValueError, TypeError):
            pass
        pri = pri_map.get(pri_val, 'P2')

        # O: 项目需求评级
        fb_id = story.get('feedback', 0)
        fb_info = None
        if fb_id and fb_id > 0:
            fb_info = feedback_dict.get(int(fb_id))
        rating = ''
        if fb_info:
            rating = fb_info.get('PRJLevel', '') or ''
            if rating in ('0', '无'):
                rating = ''

        # P: 是否对接第三方
        third_party = contains_third_party(story.get('title', ''))

        # Q: 是否为个性化
        custom = contains_custom(story.get('title', ''))

        # R: 是否为争议
        dispute = '否'

        # S: 需求名称
        req_name = story.get('title', '')

        # T: 需求提交时间
        # 优先取首次reviewed时间，其次linked2project，最后openedDate
        submit_time_str = (
            story_first_reviewed.get(story_id)
            or story_first_linked.get(story_id)
            or story.get('openedDate', '')
        )
        submit_time = parse_date(submit_time_str)

        # U: 对应研发
        dev_name = get_realname(dev_account, user_dict) if dev_account and dev_account not in ('', 'closed') else '/'

        # V: 对应测试
        qa_account = story_qa_map.get(story_id, '')
        qa_name = get_realname(qa_account, user_dict) if qa_account and qa_account not in ('', 'closed', '/') else '/'

        # W: 提测时间
        # 有task的story取最大task openedDate，无task取None
        test_time_str = story_max_task_opened.get(story_id, '')
        test_time = parse_date(test_time_str) if test_time_str else None

        # X: 验收是否通过
        story_status = story.get('status', '')
        yanshou = '通过' if story_status == 'closed' else ('待验收' if story_status == 'active' else '')

        # Y: 未通过分类
        weiguotong_fenlei = None

        # Z: 未通过原因
        weiguotong_yuanyin = None

        # AA: 是否延期提测
        yanqi = None

        # AB: 备注
        beizhu = None

        # AC: 被指派时间
        # 优先action表首次assigned，其次story.openedDate
        zhipai_str = story_first_assigned.get(story_id) or story.get('openedDate', '')
        zhipai_time = parse_date(zhipai_str)

        # AD: 响应时长（天）= 发版完成时间 - 被指派时间
        xiangying_chang = None
        if zhipai_time and faban_dt:
            diff = (faban_dt - zhipai_time).total_seconds() / 86400
            if diff >= 0:
                xiangying_chang = round(diff, 10)

        # AE: 消耗工时
        xiaohao_gongshi = story_consumed.get(story_id, None)

        # AF: 发版完成时间 → 统一使用project.end日期（模板中所有story使用相同的发版时间）
        faban_time = faban_dt  # 使用project.end

        # AG: 机构名称
        jigou_name = ''
        if fb_info:
            jigou_name = fb_info.get('organization', '') or ''
            if jigou_name in ('0',):
                jigou_name = ''

        # AH: 大区
        daqu = ''
        if fb_info:
            daqu = fb_info.get('region', '') or ''
            if daqu in ('0',):
                daqu = ''

        # AI: 分区
        fenqu = daqu

        # AJ: 反馈优先级
        fankui_pri = ''
        if fb_info:
            fp = fb_info.get('pri', '')
            if fp and fp not in (0, '0'):
                fankui_pri = str(fp)

        # AK: 反馈创建时间（新增列）
        fb_created_str = ''
        if fb_info:
            fb_created_str = fb_info.get('openedDate', '') or ''
        feedback_create_time = parse_date(fb_created_str) if fb_created_str else None

        # AL: 被指派时间（第二列，与AC相同，按模板要求）
        zhipai_time_2 = zhipai_time

        # === 新增效能指标字段 ===

        # AM: 需求估算工时 (zt_story.estimate)
        estimate_val = None
        try:
            est = story.get('estimate')
            if est is not None:
                estimate_val = float(est)
        except (ValueError, TypeError):
            pass

        # AN: 需求评审通过时间 (zt_story.reviewedDate)
        reviewed_dt = parse_date(story.get('reviewedDate', ''))

        # AO: 需求开发开始时间 (zt_task.min(realStarted))
        dev_start_str = story_min_real_started.get(story_id, '')
        dev_start_time = parse_date(dev_start_str) if dev_start_str else None

        # AP: 缺陷激活次数 (当前story的bug activatedCount)
        bug_activated = story_bug_activated.get(story_id, None)

        # AQ: 关联Bug数
        bug_cnt = story_bug_count.get(story_id, 0)

        # AR: 关联Bug激活总数
        bug_act_sum = story_bug_activated_sum.get(story_id, 0)

        # AS: 产品ID
        prod_id = story.get('product', 0)

        # AT: 产品名称
        prod_name = ''
        if prod_id and prod_id in product_dict:
            prod_name = product_dict[prod_id].get('name', '')

        # AU: 项目团队人数
        team_cnt = project_team_count.get(project_id, 0)

        # AV: 项目团队角色分布
        roles = project_team_roles.get(project_id, {})
        team_roles_str = ', '.join(f'{k}:{v}' for k, v in sorted(roles.items())) if roles else ''

        rows.append([
            month_serial,        # A  1
            week_name,           # B  2
            version_name,        # C  3
            type_name,           # D  4
            proj_id,             # E  5
            dept1,               # F  6
            dept2,               # G  7
            mod_name,            # H  8
            req_type,            # I  9
            status,              # J  10
            pm_name,             # K  11
            req_id,              # L  12
            bug_id,              # M  13
            pri,                 # N  14
            rating,              # O  15
            third_party,         # P  16
            custom,              # Q  17
            dispute,             # R  18
            req_name,            # S  19
            submit_time,         # T  20
            dev_name,            # U  21
            qa_name,             # V  22
            test_time,           # W  23
            yanshou,             # X  24
            weiguotong_fenlei,   # Y  25
            weiguotong_yuanyin,  # Z  26
            yanqi,               # AA 27
            beizhu,              # AB 28
            zhipai_time,         # AC 29
            xiangying_chang,     # AD 30
            xiaohao_gongshi,     # AE 31
            faban_time,          # AF 32
            jigou_name,          # AG 33
            daqu,                # AH 34
            fenqu,               # AI 35
            fankui_pri,          # AJ 36
            feedback_create_time,# AK 37
            zhipai_time_2,       # AL 38
            estimate_val,        # AM 39 需求估算工时
            reviewed_dt,         # AN 40 需求评审通过时间
            dev_start_time,      # AO 41 需求开发开始时间
            bug_activated,       # AP 42 缺陷激活次数
            bug_cnt,             # AQ 43 关联Bug数
            bug_act_sum,         # AR 44 关联Bug激活总数
            prod_id,             # AS 45 产品ID
            prod_name,           # AT 46 产品名称
            team_cnt,            # AU 47 项目团队人数
            team_roles_str,      # AV 48 项目团队角色分布
        ])

    print(f"✅ 组装完成：{len(rows)} 条需求记录（目标项目: {sorted(proj_ids)}）")
    return rows


def write_excel(rows, project_id=None):
    """写入Excel文件，格式对齐参考模板"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    suffix = f'project_{project_id}' if project_id else 'all'
    output_path = os.path.join(OUTPUT_DIR, f'需求跟踪矩阵_{suffix}_{timestamp}.xlsx')

    wb = Workbook()
    ws = wb.active
    ws.title = '需求跟踪矩阵'

    header_font = Font(name='微软雅黑', size=11, bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    num_cols = len(COLUMN_HEADERS)
    for col_idx, header in enumerate(COLUMN_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data_font = Font(name='微软雅黑', size=10)
    data_align = Alignment(vertical='center', wrap_text=True)

    date_cols = {20, 23, 29, 32, 37, 38, 40, 41}  # T, W, AC, AF, AK, AL, AN, AO (1-indexed)
    serial_col = 1  # A 月度
    int_cols = {5, 12, 13, 43, 44, 45, 47}  # E, L, M, AQ, AR, AS, AU
    num_cols_set = {30, 31, 39}  # AD, AE, AM (估算工时)

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

            if col_idx in date_cols and isinstance(value, datetime):
                cell.number_format = 'yyyy-mm-dd hh:mm:ss'
            elif col_idx == serial_col and isinstance(value, (int, float)):
                cell.number_format = 'yyyy/m/d'
            elif col_idx in int_cols and isinstance(value, (int, float)) and value is not None:
                cell.number_format = '0'
            elif col_idx in num_cols_set and isinstance(value, (int, float)) and value is not None:
                cell.number_format = '0.0'

    col_widths = {
        1: 8, 2: 10, 3: 28, 4: 8, 5: 10,
        6: 16, 7: 16, 8: 16, 9: 10, 10: 10,
        11: 12, 12: 10, 13: 10, 14: 8, 15: 12,
        16: 14, 17: 12, 18: 10, 19: 60, 20: 22,
        21: 12, 22: 12, 23: 22, 24: 14, 25: 12,
        26: 20, 27: 14, 28: 12, 29: 22, 30: 10,
        31: 10, 32: 22, 33: 30, 34: 12, 35: 12,
        36: 12, 37: 22, 38: 22,
        39: 12, 40: 22, 41: 22, 42: 12,
        43: 12, 44: 14, 45: 10, 46: 20,
        47: 12, 48: 30,
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(num_cols)}{len(rows)+1}'

    wb.save(output_path)
    print(f"✅ 已保存至：{output_path}")
    return output_path


def main():
    print("=" * 50)
    print("🚀 禅道数据组装脚本 (Phase 2)")
    print("=" * 50)

    # 1. 加载数据
    tables = load_data()
    print(f"📊 加载数据完成")
    for name in ['project', 'projectstory', 'story', 'task', 'bug',
                 'user', 'module', 'action', 'feedback', 'team', 'product']:
        data_rows = tables[name]
        print(f"  {name}: {len(data_rows)} 条")

    # 2. 组装矩阵（按target project过滤）
    target_ids = TARGET_PROJECT_IDS if TARGET_PROJECT_IDS else None
    rows = assemble_matrix(tables, target_project_ids=target_ids)

    # 3. 打印预览
    print("\n📋 数据预览（前3行，38列）：")
    for i in range(min(3, len(rows))):
        print(f"\n--- 第 {i+1} 行 (story_id={rows[i][11]}) ---")
        for j, (header, val) in enumerate(zip(COLUMN_HEADERS, rows[i])):
            letter = get_column_letter(j + 1)
            print(f"  {letter:3s} {header:45s}: {val}")

    # 4. 保存Excel
    proj_id = target_ids[0] if target_ids and len(target_ids) == 1 else None
    write_excel(rows, project_id=proj_id)

    print("\n🎉 数据组装完成！")


if __name__ == '__main__':
    main()
