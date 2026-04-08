#!/usr/bin/env python3
"""
业务对账自动化脚本 v3（独立版 - 完整版）
功能：
1. 从当前目录自动识别最新的源文件
2. 动态计算所有业务数据（无固定值）
3. 生成流水-YYYYMMDD.xlsx
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
import glob
import os
import re
import sys

# ==================== 辅助函数 ====================

def find_latest_file(dir_path, pattern):
    files = sorted(glob.glob(os.path.join(dir_path, pattern)), key=os.path.getmtime, reverse=True)
    return files[0] if files else None

def from_pivot(pivot_df, 业务类目):
    row = pivot_df[pivot_df['业绩类目'] == 业务类目]
    if len(row) == 0:
        return 0.0, 0
    return float(row['金额'].values[0]), int(row['订单'].values[0])

def add_multiple(items):
    if not items:
        return 0, 0
    if len(items[0]) == 3:
        return sum(float(i[1]) for i in items), sum(int(i[2]) for i in items)
    else:
        return sum(float(i[0]) for i in items), sum(int(i[1]) for i in items)

# ==================== 主程序 ====================

print('='*80)
print('📊 业务对账自动化脚本 v3')
print('='*80)

current_dir = os.getcwd()

# 自动识别文件
stats_file = find_latest_file(current_dir, '业务对账统计明细*.xlsx')
order_file = find_latest_file(current_dir, '订单明细表*.xlsx')
consult_files = glob.glob(os.path.join(current_dir, '*咨询*.xlsx'))

if not stats_file:
    print('❌ 未找到业务对账统计明细文件（需要：业务对账统计明细*.xlsx）')
    sys.exit(1)
if not order_file:
    print('❌ 未找到订单明细表文件（需要：订单明细表*.xlsx）')
    sys.exit(1)

date_match = re.search(r'(\d{8})', os.path.basename(stats_file))
today_str = date_match.group(1) if date_match else datetime.now().strftime('%Y%m%d')

print(f'\n📊 文件: {os.path.basename(stats_file)} | {os.path.basename(order_file)}')
print(f'📅 日期: {today_str}')

# 读取数据
df = pd.read_excel(stats_file)

# 筛选
df收费 = df[df['收退标识'] == '收费'].copy()
df_filtered = df收费[~df收费['机构名称'].isin(['黑龙江中医药大学附属第一医院', '上海安达医院'])].copy()
df_filtered = df_filtered[~df_filtered['业绩类目'].str.contains('医院咨询|医院复诊|医院护理', na=False)].copy()

# 透视表
pivot = pd.pivot_table(df_filtered, index='业绩类目', values='实际支付金额', aggfunc=['sum', 'count'])
pivot.columns = ['金额', '订单']
pivot = pivot.reset_index()

print('\n⏳ 读取数据...')

# ==================== 动态计算各业务 ====================

# 1. 咨询/复诊/护理
print('📊 咨询/复诊/护理...')
咨询金额, 咨询订单 = 0, 0
if consult_files:
    try:
        df_consult = pd.read_excel(consult_files[0], sheet_name='咨询')
        df_consult = df_consult[df_consult['支付状态'].isin(['已支付', '退款成功'])]
        咨询金额 = df_consult['咨询费用'].sum()
        咨询订单 = df_consult['咨询费用'].count()
    except:
        pass
自营_sum, 自营_count = from_pivot(pivot, '自营咨询/复诊/护理（医疗类相关业务）')
咨询合计 = (自营_sum + 咨询金额, 自营_count + 咨询订单)
print(f'  ✅ {咨询合计[0]:.2f} 元, {咨询合计[1]} 笔')

# 2. 处方服务
print('📊 处方服务...')
df_order = pd.read_excel(order_file)
运费 = (df_order['运费'].sum(), df_order['运费'].count())
方便购药 = from_pivot(pivot, '处方服务-便捷购药')
电子处方 = from_pivot(pivot, '处方服务-电子处方')
处方服务 = add_multiple([运费, 方便购药, 电子处方])
print(f'  ✅ {处方服务[0]:.2f} 元, {处方服务[1]} 笔')

# 3. 动态获取三家医院数据
print('📊 动态获取医院数据...')
杭州_third = df_filtered[(df_filtered['机构名称'] == '杭州市第七人民医院') & (df_filtered['业绩子类目'] == '第三方其他（病历复印等）')]
杭州_third_sum = float(杭州_third['实际支付金额'].sum()) if len(杭州_third) > 0 else 0.0
杭州_third_count = int(杭州_third['实际支付金额'].count()) if len(杭州_third) > 0 else 0

海宁_four = df_filtered[(df_filtered['机构名称'] == '海宁四院') & (df_filtered['业绩类目'] == '第三方服务')]
海宁_four_sum = float(海宁_four['实际支付金额'].sum()) if len(海宁_four) > 0 else 0.0
海宁_four_count = int(海宁_four['实际支付金额'].count()) if len(海宁_four) > 0 else 0

绍兴_three = df_filtered[(df_filtered['机构名称'] == '绍兴市第七人民医院') & (df_filtered['业绩类目'] == '第三方服务')]
绍兴_three_sum = float(绍兴_three['实际支付金额'].sum()) if len(绍兴_three) > 0 else 0.0
绍兴_three_count = int(绍兴_three['实际支付金额'].count()) if len(绍兴_three) > 0 else 0

绍兴_self = df_filtered[(df_filtered['机构名称'] == '绍兴市第七人民医院') & (df_filtered['业绩类目'] == '自营健管')]
绍兴_self_sum = float(绍兴_self['实际支付金额'].sum()) if len(绍兴_self) > 0 else 0.0
绍兴_self_count = int(绍兴_self['实际支付金额'].count()) if len(绍兴_self) > 0 else 0

# 4. 第三方服务
print('📊 第三方服务...')
第三方 = from_pivot(pivot, '第三方服务')
第三方合计 = (杭州_third_sum + 海宁_four_sum + 绍兴_three_sum, 杭州_third_count + 海宁_four_count + 绍兴_three_count)
第三方服务 = (第三方[0] - 第三方合计[0], 第三方[1] - 第三方合计[1])
print(f'  ✅ {第三方服务[0]:.2f} 元, {第三方服务[1]} 笔')

# 5. 自营健管
print('📊 自营健管...')
健管 = from_pivot(pivot, '自营健管')
自营健管 = (健管[0] - 绍兴_self_sum, 健管[1] - 绍兴_self_count)
print(f'  ✅ {自营健管[0]:.2f} 元, {自营健管[1]} 笔')

# 6. 心理咨询
print('📊 心理咨询...')
心理咨询 = (杭州_third_sum + 海宁_four_sum + 绍兴_three_sum + 绍兴_self_sum,
            杭州_third_count + 海宁_four_count + 绍兴_three_count + 绍兴_self_count)
print(f'  ✅ {心理咨询[0]:.2f} 元, {心理咨询[1]} 笔')

# 7. 其他业务
print('📊 其他业务...')
自营体检 = from_pivot(pivot, '自营体检')
会员服务 = from_pivot(pivot, '会员服务')
支付类 = from_pivot(pivot, '支付类')
print(f'  ✅ 自营体检: {自营体检[0]:.2f} 元, {自营体检[1]} 笔')
print(f'  ✅ 会员服务: {会员服务[0]:.2f} 元, {会员服务[1]} 笔')
print(f'  ✅ 支付类: {支付类[0]:.2f} 元, {支付类[1]} 笔')

# ==================== 生成流水文件 ====================

print(f'\n📝 生成流水-{today_str}.xlsx...')
output_file = os.path.join(current_dir, f'流水-{today_str}.xlsx')

wb = Workbook()
ws = wb.active

ws['A1'] = '每日新增流水详情'
ws['B2'] = '咨询/复诊/护理（医疗类相关业务）'; ws['D2'] = '处方服务\n(便捷购药)'
ws['F2'] = '自营体检'; ws['H2'] = '自营健管'; ws['J2'] = '第三方服务'
ws['L2'] = '心理咨询'; ws['N2'] = '支付类'; ws['P2'] = '远程服务'; ws['R2'] = '会员服务'

ws['C4'] = '流水'; ws['D4'] = '订单'
ws['E4'] = '流水'; ws['F4'] = '订单'
ws['G4'] = '流水'; ws['H4'] = '订单'
ws['I4'] = '流水'; ws['J4'] = '订单'
ws['K4'] = '流水'; ws['L4'] = '订单'
ws['M4'] = '流水'; ws['N4'] = '订单'
ws['O4'] = '流水'; ws['P4'] = '订单'
ws['Q4'] = '流水'; ws['R4'] = '订单'
ws['S4'] = '流水'; ws['T4'] = '订单'

ws['A5'] = datetime.strptime(today_str, '%Y%m%d')
ws['C5'], ws['D5'] = 咨询合计
ws['E5'], ws['F5'] = 处方服务
ws['G5'], ws['H5'] = 自营体检
ws['I5'], ws['J5'] = 自营健管
ws['K5'], ws['L5'] = 第三方服务
ws['M5'], ws['N5'] = 心理咨询
ws['O5'], ws['P5'] = 支付类
ws['Q5'], ws['R5'] = 会员服务

wb.save(output_file)
wb.close()

print(f'✅ 完成！已生成: {output_file}')
