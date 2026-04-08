#!/usr/bin/env python3
"""
业务对账自动化脚本（独立版）
使用方式：将此脚本复制到包含4个源文件的文件夹，直接运行
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook, Workbook
import glob
import os
import re
import shutil
import sys

# ==================== 通用函数 ====================

def from_pivot(pivot_df, 业务类目):
    """从透视表中提取单个业务类目的金额和订单数"""
    row = pivot_df[pivot_df['业绩类目'] == 业务类目]
    if len(row) == 0:
        return 0, 0
    return float(row['实际支付金额_求和'].values[0]), int(row['实际支付金额_计数'].values[0])

def deduct_hospitals(pivot_df, 业务类目, hospital_details):
    """从总数据中减去指定医院的数据"""
    total_sum, total_count = from_pivot(pivot_df, 业务类目)
    subtract_sum = sum(h[1] for h in hospital_details)
    subtract_count = sum(h[2] for h in hospital_details)
    return total_sum - subtract_sum, total_count - subtract_count

def add_multiple(items):
    """多个项目金额订单求和（兼容二元组和三元组）"""
    if not items:
        return 0, 0
    if len(items[0]) == 3:
        total_amount = sum(float(i[1]) for i in items)
        total_count = sum(int(i[2]) for i in items)
    else:
        total_amount = sum(float(i[0]) for i in items)
        total_count = sum(int(i[1]) for i in items)
    return total_amount, total_count

def find_latest_file(pattern):
    """根据模式查找最新文件"""
    files = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    return files[0] if files else None

# ==================== 主程序 ====================

print('='*80)
print('📊 业务对账自动化脚本')
print('='*80)

# 自动识别当前目录下的文件
current_dir = os.getcwd()

stats_file = find_latest_file(os.path.join(current_dir, '业务对账统计明细*.xlsx'))
if not stats_file:
    print(f'❌ 未找到业务对账统计明细文件（需要包含：业务对账统计明细*.xlsx）')
    sys.exit(1)

order_file = find_latest_file(os.path.join(current_dir, '订单明细表*.xlsx'))
if not order_file:
    print(f'❌ 未找到订单明细表文件（需要包含：订单明细表*.xlsx）')
    sys.exit(1)

# 读取日期（从文件名提取）
date_match = re.search(r'(\d{8})', os.path.basename(stats_file))
today_str = date_match.group(1) if date_match else datetime.now().strftime('%Y%m%d')

print(f'\n📊 自动识别文件：')
print(f'  业务对账统计明细: {os.path.basename(stats_file)}')
print(f'  订单明细表: {os.path.basename(order_file)}')
print(f'  日期: {today_str}')

# 读取数据
print('\n⏳ 读取数据...')
df = pd.read_excel(stats_file)

# 筛选逻辑
df收费 = df[df['收退标识'] == '收费'].copy()
exclude_hospitals = ['黑龙江中医药大学附属第一医院', '上海安达医院']
df_filtered = df收费[~df收费['机构名称'].isin(exclude_hospitals)].copy()
df_filtered = df_filtered[~df_filtered['业绩类目'].str.contains('医院咨询|医院复诊|医院护理', na=False)].copy()

# 创建透视表
pivot_table = pd.pivot_table(
    df_filtered, index='业绩类目', values='实际支付金额',
    aggfunc=['sum', 'count'], fill_value=0, margins=True, margins_name='总计'
)
pivot_table.columns = ['实际支付金额_求和', '实际支付金额_计数']
pivot_table = pivot_table.reset_index()

# ==================== 业务计算 ====================

print('📊 计算各业务模块...')

# 1. 咨询/复诊/护理
consult_file_pattern = os.path.join(current_dir, '*咨询*.xlsx')
consult_files = glob.glob(consult_file_pattern)
咨询金额, 咨询订单 = 0, 0
if consult_files:
    try:
        df_consult = pd.read_excel(consult_files[0], sheet_name='咨询')
        df_consult = df_consult[df_consult['支付状态'].isin(['已支付', '退款成功'])]
        咨询金额 = df_consult['咨询费用'].sum()
        咨询订单 = df_consult['咨询费用'].count()
    except:
        pass

自营_sum, 自营_count = from_pivot(pivot_table, '自营咨询/复诊/护理（医疗类相关业务）')
咨询合计 = (自营_sum + 咨询金额, 自营_count + 咨询订单)
print(f'  ✅ 咨询/复诊/护理: {咨询合计[0]:.2f} 元, {咨询合计[1]} 笔')

# 2. 处方服务
df_order = pd.read_excel(order_file)
运费 = (df_order['运费'].sum(), df_order['运费'].count())
方便购药 = from_pivot(pivot_table, '处方服务-便捷购药')
电子处方 = from_pivot(pivot_table, '处方服务-电子处方')
处方服务 = add_multiple([运费, 方便购药, 电子处方])
print(f'  ✅ 处方服务: {处方服务[0]:.2f} 元, {处方服务[1]} 笔')

# 3. 第三方服务（减去三家医院）
第三方 = from_pivot(pivot_table, '第三方服务')
三家 = [('杭州第七医院', 30214.40, 97), ('海宁四院', 2472.01, 10), ('绍兴七院', 300.00, 1)]
第三方服务 = deduct_hospitals(pivot_table, '第三方服务', 三家)
print(f'  ✅ 第三方服务: {第三方服务[0]:.2f} 元, {第三方服务[1]} 笔')

# 4. 自营健管（减去绍兴七院）
健管 = from_pivot(pivot_table, '自营健管')
绍兴 = (11700.00, 29)
自营健管 = (健管[0] - 绍兴[0], 健管[1] - 绍兴[1])
print(f'  ✅ 自营健管: {自营健管[0]:.2f} 元, {自营健管[1]} 笔')

# 5. 心理咨询（4项相加）
心理 = [
    ('杭州第七医院-第三方', 30214.40, 97),
    ('海宁四院-第三方', 2472.01, 10),
    ('绍兴七院-第三方', 300.00, 1),
    ('绍兴七院-自营健管', 11700.00, 29)
]
心理咨询 = add_multiple(心理)
print(f'  ✅ 心理咨询: {心理咨询[0]:.2f} 元, {心理咨询[1]} 笔')

# 6. 其他业务（从透视表动态计算）
体检_sum, 体检_count = from_pivot(pivot_table, '自营体检')
print(f'  ✅ 自营体检: {体检_sum:.2f} 元, {体检_count} 笔')

会员_sum, 会员_count = from_pivot(pivot_table, '会员服务')
print(f'  ✅ 会员服务: {会员_sum:.2f} 元, {会员_count} 笔')

支付_sum, 支付_count = from_pivot(pivot_table, '支付类')
print(f'  ✅ 支付类: {支付_sum:.2f} 元, {支付_count} 笔')

# ==================== 生成流水文件 ====================

print(f'\n📝 生成流水-{today_str}.xlsx...')
output_file = os.path.join(current_dir, f'流水-{today_str}.xlsx')

# 使用 openpyxl 创建与流水test.xlsx相同格式的文件
wb = Workbook()
ws = wb.active

# 行1: 标题
ws['A1'] = '每日新增流水详情'

# 行2: 业务列名（按原格式）
ws['B2'] = '咨询/复诊/护理（医疗类相关业务）'
ws['D2'] = '处方服务\n(便捷购药)'
ws['F2'] = '自营体检'
ws['H2'] = '自营健管'
ws['J2'] = '第三方服务'
ws['L2'] = '心理咨询'
ws['N2'] = '支付类'
ws['P2'] = '会员服务'

# 行3: 空行

# 行4: 列标题（流水/订单交替）
ws['C4'] = '流水'
ws['D4'] = '订单'
ws['E4'] = '流水'
ws['F4'] = '订单'
ws['G4'] = '流水'
ws['H4'] = '订单'
ws['I4'] = '流水'
ws['J4'] = '订单'
ws['K4'] = '流水'
ws['L4'] = '订单'
ws['M4'] = '流水'
ws['N4'] = '订单'
ws['O4'] = '流水'
ws['P4'] = '订单'

# 行5: 数据
ws['A5'] = datetime.strptime(today_str, '%Y%m%d')
ws['C5'] = 咨询合计[0]
ws['D5'] = 咨询合计[1]
ws['E5'] = 处方服务[0]
ws['F5'] = 处方服务[1]
ws['G5'] = 自营体检[0]
ws['H5'] = 自营体检[1]
ws['I5'] = 自营健管[0]
ws['J5'] = 自营健管[1]
ws['K5'] = 第三方服务[0]
ws['L5'] = 第三方服务[1]
ws['M5'] = 心理咨询[0]
ws['N5'] = 心理咨询[1]
ws['O5'] = 支付类[0]
ws['P5'] = 支付类[1]
# 会员服务在更后面，先留空

# 保存
wb.save(output_file)
wb.close()

print(f'✅ 完成！已生成: {output_file}')
