#!/usr/bin/env python3
"""
禅道需求+Bug数据库一键导入脚本
从Excel文件读取数据，写入SQLite数据库
"""

import os
import re
import sys
import sqlite3
import logging
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

# ============================================================
# 配置
# ============================================================
BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = BASE_DIR / "data"
DB_PATH = DATA_DIR / "zentao_dashboard.db"
SCHEMA_PATH = BASE_DIR / "db_schema.sql"

REQ_DIR = Path("/mnt/d/钉钉下载/1研发中心质控部基础数据/需求表")
BUG_DIR = Path("/mnt/d/钉钉下载/1研发中心质控部基础数据/bug表")

REQ_FILES = [
    REQ_DIR / "2026-1-V1版本-01-需求跟踪矩阵.xlsx",
    REQ_DIR / "2026-1-V2版本-01-需求跟踪矩阵.xlsx",
    REQ_DIR / "2026-2-V1版本-01-需求跟踪矩阵.xlsx",
    REQ_DIR / "2026-3-V1版本-01-需求跟踪矩阵.xlsx",
    REQ_DIR / "2026-3-V2版本-01-需求跟踪矩阵.xlsx",
]

BUG_FILES = [
    BUG_DIR / "2026-1-V1版本-Bug总数及分布.xlsx",
    BUG_DIR / "2026-1-V2版本-Bug总数及分布.xlsx",
    BUG_DIR / "2026-2-V1版本-Bug总数及分布.xlsx",
    BUG_DIR / "2026-3-V1版本-Bug总数及分布.xlsx",
    BUG_DIR / "2026-4-V1版本-Bug总数及分布.xlsx",
    BUG_DIR / "2026-4-V2版本-Bug总数及分布.xlsx",
]

# ============================================================
# 日志
# ============================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger("zentao_import")

# ============================================================
# 工具函数
# ============================================================

def excel_serial_to_date(serial):
    """Excel序列号转日期字符串 YYYY-MM-DD HH:MM:SS"""
    if serial is None:
        return None
    if isinstance(serial, (int, float)):
        if serial < 1:
            return None
        base = datetime(1899, 12, 30)
        # Excel有个bug：把1900-02-29当作有效日期
        days = int(serial)
        frac = serial - days
        dt = base + timedelta(days=days)
        if frac > 0:
            hours = int(frac * 24)
            minutes = int((frac * 24 - hours) * 60)
            seconds = int(((frac * 24 - hours) * 60 - minutes) * 60)
            dt = dt.replace(hour=hours, minute=minutes, second=seconds)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    return None


def parse_date(value):
    """统一日期解析：datetime对象 / Excel序列号 / 字符串"""
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, (int, float)):
        return excel_serial_to_date(value)

    if isinstance(value, str):
        s = value.strip()
        if s in ("无", "/", "", "#N/A", "0000-00-00", "0000-00-00 00:00:00"):
            return None
        # 尝试解析标准日期格式
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d"):
            try:
                return datetime.strptime(s, fmt).strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
        # 尝试Excel序列号
        try:
            return excel_serial_to_date(float(s))
        except (ValueError, TypeError):
            return s
    return None


def clean_text(value):
    """清理文本字段"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return str(value)
    s = str(value).strip()
    if s in ("", "/", "#N/A", "无"):
        return None
    return s


def clean_req_id(value):
    """需求ID: float → string，去掉 .0 后缀"""
    if value is None:
        return None
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value)
    return str(value).strip()


def extract_version_code(filename):
    """从文件名提取版本编码，如 2026-1-V1"""
    m = re.match(r"(\d{4}-\d+-V\d+)", Path(filename).stem)
    return m.group(1) if m else None


def extract_version_name(version_name_raw):
    """从版本名称字符串提取，如 '2026-1-V1版本（0112）'"""
    if version_name_raw:
        return str(version_name_raw).strip()
    return None


def extract_release_date(version_name_raw):
    """从版本名称中提取发版日期，如 '2026-1-V1版本（0112）' → 2026-01-12"""
    if not version_name_raw:
        return None
    m = re.search(r"（(\d{2})(\d{2})）", str(version_name_raw))
    if m:
        # 年份从版本编码推断，但这里只有月和日
        # 例如 0112 → 01-12，需要从版本编码拿年份
        return f"{m.group(1)}-{m.group(2)}"
    return None


def clean_person_name(name):
    """去掉人员名称后缀: -产品经理/-开发/-测试"""
    if not name:
        return None
    s = str(name).strip()
    if s in ("", "/", "#N/A"):
        return None
    # 去掉后缀
    s = re.sub(r"-产品经理$|-开发$|-测试$", "", s)
    return s if s else None


def normalize_dept(value):
    """部门名称统一"""
    if value is None:
        return None
    s = str(value).strip()
    if s == "产品&数据中心":
        return "产品与数据中心"
    if s in ("", "/", "#N/A"):
        return None
    return s


def safe_int(value, default=None):
    """安全转整数"""
    if value is None:
        return default
    try:
        return int(float(str(value).strip()))
    except (ValueError, TypeError):
        return default


def safe_float(value, default=None):
    """安全转浮点数"""
    if value is None:
        return default
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return default


# ============================================================
# 初始化数据库
# ============================================================

def init_db(db_path):
    """创建数据库和表结构"""
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    # 读取并执行schema
    schema_sql = open(SCHEMA_PATH, "r", encoding="utf-8").read()
    conn.executescript(schema_sql)
    logger.info(f"数据库已初始化: {db_path}")
    return conn


# ============================================================
# 导入需求数据
# ============================================================

def find_req_sheet(wb):
    """找到需求跟踪矩阵sheet"""
    for name in wb.sheetnames:
        if "需求跟踪矩阵" in name:
            return name
    return None


def find_change_log_sheet(wb):
    """找到需求变更记录sheet"""
    for name in wb.sheetnames:
        if "需求变更" in name:
            return name
    return None


def find_staff_sheet(wb):
    """找到人员信息sheet"""
    for name in wb.sheetnames:
        if "人员信息" in name:
            return name
    return None


def find_hours_sheet(wb):
    """找到Sheet2（工时数据）"""
    for name in wb.sheetnames:
        if name == "Sheet2":
            return name
    return None


def import_requirements(conn, filepath):
    """导入单个需求文件"""
    version_code = extract_version_code(filepath)
    logger.info(f"导入需求: {filepath.name} (版本: {version_code})")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 1. 先处理人员信息
    staff_sheet = find_staff_sheet(wb)
    if staff_sheet:
        import_staff_from_file(conn, wb, staff_sheet, version_code)

    # 2. 找需求跟踪矩阵sheet
    req_sheet = find_req_sheet(wb)
    if not req_sheet:
        logger.warning(f"未找到需求跟踪矩阵sheet: {filepath.name}")
        wb.close()
        return 0

    ws = wb[req_sheet]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_count = len(headers)
    logger.info(f"需求sheet: {req_sheet}, 列数={col_count}")

    # 从版本名称行提取信息
    sample_row = ws.cell(row=2, column=3).value  # C列 = 版本名称
    version_name = extract_version_name(sample_row)
    release_date_partial = extract_release_date(sample_row)
    sample_type = clean_text(ws.cell(row=2, column=4).value)
    sample_project_id = safe_int(ws.cell(row=2, column=5).value)

    # 计算release_date (需要年份)
    release_date = None
    if version_code and release_date_partial:
        year = version_code.split("-")[0]
        release_date = f"{year}-{release_date_partial}"

    # 插入或获取version记录
    cur = conn.cursor()
    cur.execute("SELECT id FROM version WHERE version_code = ?", (version_code,))
    row = cur.fetchone()
    if row:
        version_id = row[0]
        logger.info(f"版本 {version_code} 已存在 (id={version_id})，跳过")
        wb.close()
        return 0
    else:
        cur.execute(
            "INSERT INTO version (version_code, version_name, release_date, version_type, project_id, source_file) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (version_code, version_name, release_date, sample_type, sample_project_id, str(filepath))
        )
        version_id = cur.lastrowid

    # 3. 读取需求数据 (从第2行开始)
    success = 0
    failed = 0
    total = 0

    # 列索引映射 (0-based)
    # A:月度(0), B:周数(1), C:版本名称(2), D:类型(3), E:项目ID(4),
    # F:一级部门(5), G:二级部门(6), H:所属业务模块(7), I:需求类型(8), J:需求状态(9),
    # K:产品经理(10), L:需求ID(11), M:bugID(12), N:优先级(13), O:项目需求评级(14),
    # P:是否对接第三方(15), Q:是否为个性化(16), R:是否为争议(17), S:需求名称(18),
    # T:需求提交时间(19), U:对应研发(20), V:对应测试(21), W:提测时间(22),
    # X:验收是否通过(23), Y:未通过分类(24), Z:未通过原因(25), AA:是否延期提测(26),
    # AB:备注(27), AC:被指派时间(28), AD:响应时长(29), AE:消耗工时(30),
    # AF:发版完成时间(31), AG:机构名称(32), AH:大区(33), AI:分区(34),
    # AJ:反馈优先级(35), AK:反馈创建时间(36)
    # 某些文件有额外列: AL:被指派时间(37), AM:项目需求紧迫度(38), AN:成本(39)

    for r in range(2, ws.max_row + 1):
        req_id_raw = ws.cell(row=r, column=12).value  # L列
        if req_id_raw is None:
            continue
        # 跳过空行或标题行
        req_id = clean_req_id(req_id_raw)
        if not req_id or req_id in ("需求ID", "需求编号"):
            continue

        total += 1
        try:
            req_title = clean_text(ws.cell(row=r, column=19).value)
            req_type = clean_text(ws.cell(row=r, column=9).value)
            req_status = clean_text(ws.cell(row=r, column=10).value)
            priority = clean_text(ws.cell(row=r, column=14).value)
            project_rating = clean_text(ws.cell(row=r, column=15).value)
            dept_level1 = normalize_dept(ws.cell(row=r, column=6).value)
            dept_level2 = normalize_dept(ws.cell(row=r, column=7).value)
            business_module = clean_text(ws.cell(row=r, column=8).value)
            product_manager = clean_person_name(ws.cell(row=r, column=11).value)
            developer = clean_person_name(ws.cell(row=r, column=21).value)
            tester = clean_person_name(ws.cell(row=r, column=22).value)

            submit_time = parse_date(ws.cell(row=r, column=20).value)
            test_submit_time = parse_date(ws.cell(row=r, column=23).value)
            release_time = parse_date(ws.cell(row=r, column=32).value)

            consumed_hours = safe_float(ws.cell(row=r, column=31).value)

            is_third_party = clean_text(ws.cell(row=r, column=16).value)
            is_customized = clean_text(ws.cell(row=r, column=17).value)
            is_disputed = clean_text(ws.cell(row=r, column=18).value)
            is_delayed_test = clean_text(ws.cell(row=r, column=27).value)
            acceptance_status = clean_text(ws.cell(row=r, column=24).value)

            region = clean_text(ws.cell(row=r, column=34).value)
            sub_region = clean_text(ws.cell(row=r, column=35).value)
            org_name = clean_text(ws.cell(row=r, column=33).value)
            feedback_priority = clean_text(ws.cell(row=r, column=36).value)
            feedback_time = parse_date(ws.cell(row=r, column=37).value)
            remark = clean_text(ws.cell(row=r, column=28).value)

            cur.execute(
                """INSERT INTO requirement (
                    version_id, req_id, req_title, req_type, req_status, priority,
                    project_rating, dept_level1, dept_level2, business_module,
                    product_manager, developer, tester, submit_time, test_submit_time,
                    release_time, consumed_hours, is_third_party, is_customized,
                    is_disputed, is_delayed_test, acceptance_status, region, sub_region,
                    org_name, feedback_priority, feedback_time, remark
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    version_id, req_id, req_title, req_type, req_status, priority,
                    project_rating, dept_level1, dept_level2, business_module,
                    product_manager, developer, tester, submit_time, test_submit_time,
                    release_time, consumed_hours, is_third_party, is_customized,
                    is_disputed, is_delayed_test, acceptance_status, region, sub_region,
                    org_name, feedback_priority, feedback_time, remark
                )
            )
            success += 1
        except Exception as e:
            failed += 1
            logger.warning(f"  需求行 {r} 导入失败: {e}")

    # 4. 导入需求变更日志
    change_sheet = find_change_log_sheet(wb)
    change_success = 0
    if change_sheet:
        change_success = import_change_log(conn, wb, change_sheet, version_id)

    # 5. 导入工时数据 (Sheet2)
    hours_success = 0
    hours_sheet = find_hours_sheet(wb)
    if hours_sheet:
        hours_success = import_hours_data(conn, wb, hours_sheet, version_id, version_code)

    # 审计记录
    log_audit(conn, version_code, "requirement", str(filepath),
              "success" if failed == 0 else "partial",
              total, success, 0, failed, None)

    wb.close()
    logger.info(f"需求导入完成: 总计={total}, 成功={success}, 失败={failed}, 变更={change_success}, 工时={hours_success}")
    return success


def import_change_log(conn, wb, sheet_name, version_id):
    """导入需求变更日志"""
    ws = wb[sheet_name]
    success = 0
    cur = conn.cursor()

    for r in range(3, ws.max_row + 1):
        # 列: 版本阶段(0), 日期(1), 时间(2), 所属一级部门(3), 变更人(4),
        # 变更类型(5), 需求ID(6), 需求类型(7), 备注(8), 检查已完成(9)
        req_id_raw = ws.cell(row=r, column=7).value
        if req_id_raw is None:
            continue
        req_id = clean_req_id(req_id_raw)
        if not req_id:
            continue

        stage = clean_text(ws.cell(row=r, column=1).value)
        change_date = parse_date(ws.cell(row=r, column=2).value)
        change_time_val = ws.cell(row=r, column=3).value
        change_time = None
        if change_time_val:
            if isinstance(change_time_val, str):
                change_time = change_time_val
            elif hasattr(change_time_val, 'strftime'):
                change_time = change_time_val.strftime("%H:%M:%S")

        dept_level1 = normalize_dept(ws.cell(row=r, column=4).value)
        changer = clean_person_name(ws.cell(row=r, column=5).value)
        change_type = clean_text(ws.cell(row=r, column=6).value)
        req_type = clean_text(ws.cell(row=r, column=8).value)
        remark = clean_text(ws.cell(row=r, column=9).value)
        is_checked = safe_int(ws.cell(row=r, column=10).value)

        try:
            cur.execute(
                """INSERT INTO requirement_change_log (
                    version_id, stage, change_date, change_time, dept_level1,
                    changer, change_type, req_id, req_type, remark, is_checked
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (version_id, stage, change_date, change_time, dept_level1,
                 changer, change_type, req_id, req_type, remark, is_checked)
            )
            success += 1
        except Exception as e:
            logger.warning(f"  变更日志行 {r} 导入失败: {e}")

    return success


def import_hours_data(conn, wb, sheet_name, version_id, version_code):
    """从Sheet2导入工时和反馈数据"""
    ws = wb[sheet_name]
    success = 0
    cur = conn.cursor()

    # Sheet2 列: 项目名称(0), 需求ID(1), 关联任务数(2), 预估工时(3), 实际消耗(4), 比较(5),
    # 需求ID(6)~... 然后后面: 需求ID(14), 反馈ID(15), 评级(16), 机构名称(17), 大区(18), 分区(19), 反馈优先级(20), 反馈创建时间(21)

    for r in range(2, ws.max_row + 1):
        req_id_raw = ws.cell(row=r, column=2).value  # B列
        if req_id_raw is None:
            continue
        req_id = clean_req_id(req_id_raw)
        if not req_id:
            continue

        task_count = safe_int(ws.cell(row=r, column=3).value)
        estimated_hours = safe_float(ws.cell(row=r, column=4).value)
        actual_hours = safe_float(ws.cell(row=r, column=5).value)
        hour_diff = None
        if estimated_hours is not None and actual_hours is not None:
            hour_diff = round(actual_hours - estimated_hours, 2)

        feedback_id = clean_req_id(ws.cell(row=r, column=16).value)  # P列
        feedback_create_time = parse_date(ws.cell(row=r, column=22).value)  # V列

        # 查找对应需求ID
        cur.execute("SELECT id FROM requirement WHERE version_id = ? AND req_id = ?",
                    (version_id, req_id))
        req_row = cur.fetchone()
        req_db_id = req_row[0] if req_row else None

        try:
            cur.execute(
                """INSERT INTO requirement_hours (
                    requirement_id, task_count, estimated_hours, actual_hours,
                    hour_diff, feedback_id, feedback_create_time
                ) VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (req_db_id, task_count, estimated_hours, actual_hours,
                 hour_diff, feedback_id, feedback_create_time)
            )
            success += 1
        except Exception as e:
            logger.warning(f"  工时行 {r} 导入失败: {e}")

    return success


def import_staff_from_file(conn, wb, sheet_name, version_code):
    """从Excel人员信息sheet导入staff表"""
    ws = wb[sheet_name]
    success = 0
    cur = conn.cursor()

    # 查找表头行
    header_row = None
    for r in range(1, min(10, ws.max_row + 1)):
        val = ws.cell(row=r, column=2).value
        if val and "人员信息表" in str(val):
            header_row = r + 1  # 下一行是表头
            break

    if header_row is None:
        return 0

    for r in range(header_row + 1, ws.max_row + 1):
        name = clean_text(ws.cell(row=r, column=3).value)
        if not name:
            continue

        dept_level1 = normalize_dept(ws.cell(row=r, column=4).value)
        dept_level2 = normalize_dept(ws.cell(row=r, column=5).value)
        business_line = clean_text(ws.cell(row=r, column=6).value)
        role = clean_text(ws.cell(row=r, column=7).value)
        name_clean = clean_person_name(name)

        if business_line == "0":
            business_line = None

        try:
            cur.execute(
                "INSERT OR IGNORE INTO staff (name, name_clean, dept_level1, dept_level2, business_line, role) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (name, name_clean, dept_level1, dept_level2, business_line, role)
            )
            if cur.rowcount > 0:
                success += 1
        except Exception as e:
            pass  # unique constraint skips

    return success


# ============================================================
# 导入Bug数据
# ============================================================

def find_bug_sheet(wb):
    """动态查找原始bug sheet"""
    for name in wb.sheetnames:
        if "原始bug" in name:
            return name
    return None


def import_bugs(conn, filepath):
    """导入单个Bug文件"""
    version_code = extract_version_code(filepath)
    logger.info(f"导入Bug: {filepath.name} (版本: {version_code})")

    wb = openpyxl.load_workbook(filepath, data_only=True)

    # 1. 先处理人员信息
    staff_sheet = find_staff_sheet(wb)
    if staff_sheet:
        import_staff_from_file(conn, wb, staff_sheet, version_code)

    # 2. 找原始bug sheet
    bug_sheet = find_bug_sheet(wb)
    if not bug_sheet:
        logger.warning(f"未找到原始bug sheet: {filepath.name}")
        wb.close()
        return 0

    ws = wb[bug_sheet]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_count = len(headers)
    logger.info(f"Bug sheet: {bug_sheet}, 列数={col_count}")

    # 列索引映射 (0-based, 40列):
    # 0:测试阶段, 1:Bug编号, 2:所属产品, 3:所属模块, 4:所属项目, 5:相关需求, 6:相关任务,
    # 7:Bug标题, 8:关键词, 9:严重程度, 10:优先级, 11:Bug类型, 12:操作系统, 13:浏览器,
    # 14:Bug状态, 15:截止日期, 16:激活次数, 17:是否确认, 18:抄送给, 19:由谁创建,
    # 20:创建日期, 21:影响版本, 22:指派给, 23:指派日期, 24:解决者, 25:解决方案,
    # 26:解决版本, 27:解决日期, 28:由谁关闭, 29:关闭日期, 30:重复ID, 31:相关Bug,
    # 32:相关用例, 33:最后修改者, 34:修改日期, 35:附件, 36:所属岗位, 37:所属二级部门,
    # 38:所属一级部门, 39:所属业务线

    # 获取版本信息
    sample_project = clean_text(ws.cell(row=2, column=5).value)
    version_name = sample_project

    cur = conn.cursor()
    cur.execute("SELECT id FROM version WHERE version_code = ?", (version_code,))
    row = cur.fetchone()
    if not row:
        # 如果version不存在，创建一个
        release_date = None
        year = version_code.split("-")[0]
        parts = version_code.split("-")
        if len(parts) >= 2:
            # 尝试从文件名推断
            pass
        cur.execute(
            "INSERT INTO version (version_code, version_name, source_file) VALUES (?, ?, ?)",
            (version_code, version_name, str(filepath))
        )
        version_id = cur.lastrowid
    else:
        version_id = row[0]

    # 读取Bug数据
    success = 0
    failed = 0
    total = 0

    for r in range(2, ws.max_row + 1):
        bug_id_raw = ws.cell(row=r, column=2).value  # B列
        if bug_id_raw is None:
            continue
        bug_id = safe_int(bug_id_raw)
        if bug_id is None:
            continue

        # 检查bug标题是否为空（空行）
        bug_title = clean_text(ws.cell(row=r, column=8).value)
        if not bug_title:
            continue

        total += 1
        try:
            test_stage = clean_text(ws.cell(row=r, column=1).value)
            severity = safe_int(ws.cell(row=r, column=10).value)
            priority = safe_int(ws.cell(row=r, column=11).value)
            bug_type = clean_text(ws.cell(row=r, column=12).value)
            bug_status = clean_text(ws.cell(row=r, column=15).value)
            is_confirmed = clean_text(ws.cell(row=r, column=18).value)
            activate_count = safe_int(ws.cell(row=r, column=17).value)
            solution = clean_text(ws.cell(row=r, column=26).value)
            creator = clean_person_name(ws.cell(row=r, column=20).value)
            assignee = clean_person_name(ws.cell(row=r, column=23).value)
            resolver = clean_person_name(ws.cell(row=r, column=25).value)
            closer = clean_person_name(ws.cell(row=r, column=29).value)
            last_modifier = clean_person_name(ws.cell(row=r, column=34).value)

            create_time = parse_date(ws.cell(row=r, column=21).value)
            assign_date = parse_date(ws.cell(row=r, column=24).value)
            resolve_date = parse_date(ws.cell(row=r, column=28).value)
            close_date_raw = ws.cell(row=r, column=30).value
            close_date = parse_date(close_date_raw)
            modify_time = parse_date(ws.cell(row=r, column=35).value)

            module_path = clean_text(ws.cell(row=r, column=4).value)
            product_name = clean_text(ws.cell(row=r, column=3).value)
            project_name = clean_text(ws.cell(row=r, column=5).value)
            impact_version = clean_text(ws.cell(row=r, column=22).value)
            resolve_version = clean_text(ws.cell(row=r, column=27).value)
            duplicate_id = clean_req_id(ws.cell(row=r, column=31).value)

            has_attachment = 1 if ws.cell(row=r, column=36).value else 0

            dept_level1 = normalize_dept(ws.cell(row=r, column=39).value)
            dept_level2 = normalize_dept(ws.cell(row=r, column=38).value)
            role = clean_text(ws.cell(row=r, column=37).value)
            business_line = clean_text(ws.cell(row=r, column=40).value)

            if business_line == "0":
                business_line = None

            cur.execute(
                """INSERT OR IGNORE INTO bug (
                    version_id, bug_id, test_stage, bug_title, severity, priority,
                    bug_type, bug_status, is_confirmed, activate_count, solution,
                    creator, assignee, resolver, closer, last_modifier,
                    create_time, assign_date, resolve_date, close_date, modify_time,
                    module_path, product_name, project_name, impact_version,
                    resolve_version, duplicate_id, has_attachment,
                    dept_level1, dept_level2, role, business_line
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    version_id, bug_id, test_stage, bug_title, severity, priority,
                    bug_type, bug_status, is_confirmed, activate_count, solution,
                    creator, assignee, resolver, closer, last_modifier,
                    create_time, assign_date, resolve_date, close_date, modify_time,
                    module_path, product_name, project_name, impact_version,
                    resolve_version, duplicate_id, has_attachment,
                    dept_level1, dept_level2, role, business_line
                )
            )
            if cur.rowcount > 0:
                success += 1
            else:
                # 已存在，跳过
                pass
                success += 1  # 也算成功
        except Exception as e:
            failed += 1
            logger.warning(f"  Bug行 {r} 导入失败: {e}")

    # 审计记录
    log_audit(conn, version_code, "bug", str(filepath),
              "success" if failed == 0 else "partial",
              total, success, 0, failed, None)

    wb.close()
    logger.info(f"Bug导入完成: 总计={total}, 成功={success}, 失败={failed}")
    return success


# ============================================================
# 审计日志
# ============================================================

def log_audit(conn, version_code, source_type, source_file, status,
              total, success, skipped, failed, error_details):
    """记录导入审计"""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute(
        """INSERT INTO import_audit (version_code, source_type, source_file, import_status,
           total_rows, success_rows, skipped_rows, failed_rows, error_details, started_at, completed_at)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (version_code, source_type, source_file, status, total, success, skipped, failed,
         error_details, now, now)
    )


# ============================================================
# 统计报告
# ============================================================

def print_report(conn):
    """输出完整统计报告"""
    cur = conn.cursor()
    print("\n" + "=" * 70)
    print("  禅道 Dashboard 数据导入统计报告")
    print("=" * 70)

    # 版本总览
    print("\n📦 版本总览:")
    print("-" * 70)
    cur.execute("SELECT version_code, version_name, release_date, status, source_file FROM version ORDER BY version_code")
    for row in cur.fetchall():
        code, name, date, status, src = row
        src_name = Path(src).name if src else "N/A"
        print(f"  {code:20s} | {str(name)[:30]:30s} | {date or 'N/A':12s} | {status:8s} | {src_name}")

    # 需求统计
    print("\n📋 需求统计 (按版本):")
    print("-" * 70)
    cur.execute("""
        SELECT v.version_code, COUNT(*) AS req_count,
               COALESCE(SUM(r.consumed_hours), 0) AS total_hours,
               COUNT(DISTINCT r.dept_level1) AS dept_count
        FROM version v
        LEFT JOIN requirement r ON v.id = r.version_id
        GROUP BY v.id
        ORDER BY v.version_code
    """)
    for row in cur.fetchall():
        code, count, hours, depts = row
        print(f"  {code:20s} | 需求数: {count:4d} | 总工时: {hours:8.1f} | 涉及部门: {depts}")

    # Bug统计
    print("\n🐛 Bug统计 (按版本):")
    print("-" * 70)
    cur.execute("""
        SELECT v.version_code, COUNT(*) AS bug_count,
               SUM(CASE WHEN b.severity = 1 THEN 1 ELSE 0 END) AS sev1,
               SUM(CASE WHEN b.severity = 2 THEN 1 ELSE 0 END) AS sev2,
               SUM(CASE WHEN b.severity = 3 THEN 1 ELSE 0 END) AS sev3,
               SUM(CASE WHEN b.severity = 4 THEN 1 ELSE 0 END) AS sev4
        FROM version v
        LEFT JOIN bug b ON v.id = b.version_id
        GROUP BY v.id
        ORDER BY v.version_code
    """)
    for row in cur.fetchall():
        code, count, s1, s2, s3, s4 = row
        print(f"  {code:20s} | Bug数: {count:4d} | S1:{s1:3d} S2:{s2:3d} S3:{s3:3d} S4:{s4:3d}")

    # 需求类型分布
    print("\n📊 需求类型分布:")
    print("-" * 70)
    cur.execute("""
        SELECT req_type, COUNT(*) AS cnt,
               ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM requirement), 1) AS pct
        FROM requirement
        WHERE req_type IS NOT NULL
        GROUP BY req_type
        ORDER BY cnt DESC
    """)
    for row in cur.fetchall():
        rtype, cnt, pct = row
        print(f"  {rtype:15s} | {cnt:4d} ({pct}%)")

    # Bug解决方案分布
    print("\n🔧 Bug解决方案分布:")
    print("-" * 70)
    cur.execute("""
        SELECT solution, COUNT(*) AS cnt,
               ROUND(COUNT(*) * 100.0 / (SELECT COUNT(*) FROM bug), 1) AS pct
        FROM bug
        WHERE solution IS NOT NULL
        GROUP BY solution
        ORDER BY cnt DESC
    """)
    for row in cur.fetchall():
        sol, cnt, pct = row
        print(f"  {sol:15s} | {cnt:4d} ({pct}%)")

    # 部门需求统计
    print("\n🏢 部门需求统计:")
    print("-" * 70)
    cur.execute("""
        SELECT dept_level1, COUNT(*) AS cnt,
               ROUND(COALESCE(SUM(consumed_hours), 0), 1) AS hours
        FROM requirement
        WHERE dept_level1 IS NOT NULL
        GROUP BY dept_level1
        ORDER BY cnt DESC
    """)
    for row in cur.fetchall():
        dept, cnt, hours = row
        print(f"  {dept:20s} | 需求数: {cnt:4d} | 工时: {hours:8.1f}")

    # 人员统计
    print("\n👥 人员信息:")
    print("-" * 70)
    cur.execute("SELECT COUNT(*) FROM staff")
    print(f"  总人员数: {cur.fetchone()[0]}")
    cur.execute("""
        SELECT role, COUNT(*) AS cnt
        FROM staff
        WHERE role IS NOT NULL
        GROUP BY role
        ORDER BY cnt DESC
    """)
    for row in cur.fetchall():
        role, cnt = row
        print(f"  {role:15s} | {cnt:3d} 人")

    # 变更日志
    print("\n📝 需求变更日志:")
    print("-" * 70)
    cur.execute("SELECT COUNT(*) FROM requirement_change_log")
    print(f"  总变更记录数: {cur.fetchone()[0]}")

    # 工时数据
    print("\n⏱ 工时数据:")
    print("-" * 70)
    cur.execute("SELECT COUNT(*) FROM requirement_hours")
    print(f"  总工时记录数: {cur.fetchone()[0]}")

    # 导入审计
    print("\n🔍 导入审计记录:")
    print("-" * 70)
    cur.execute("""
        SELECT version_code, source_type, source_file, import_status,
               total_rows, success_rows, failed_rows
        FROM import_audit
        ORDER BY id
    """)
    for row in cur.fetchall():
        code, stype, src, status, total, ok, fail = row
        src_name = Path(src).name if src else "N/A"
        print(f"  {code:20s} | {stype:12s} | {status:8s} | 总计:{total:4d} 成功:{ok:4d} 失败:{fail:3d} | {src_name}")

    # 汇总
    print("\n📈 总体汇总:")
    print("-" * 70)
    cur.execute("SELECT COUNT(*) FROM version")
    print(f"  版本数: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM requirement")
    print(f"  需求总数: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM bug")
    print(f"  Bug总数: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM staff")
    print(f"  人员总数: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM requirement_change_log")
    print(f"  变更日志: {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(*) FROM requirement_hours")
    print(f"  工时记录: {cur.fetchone()[0]}")

    # View验证
    print("\n✅ VIEW 验证:")
    print("-" * 70)
    views = [
        "v_req_summary", "v_bug_summary", "v_workload", "v_version_compare",
        "v_module_distribution", "v_bug_quality", "v_req_status_flow",
        "v_bug_severity_trend", "v_req_bug_ratio_by_dept", "v_staff_workload_detail",
        "v_delay_analysis", "v_version_health_score", "v_monthly_trend"
    ]
    for view in views:
        try:
            cur.execute(f"SELECT COUNT(*) FROM {view}")
            cnt = cur.fetchone()[0]
            print(f"  ✅ {view:35s} | 记录数: {cnt}")
        except Exception as e:
            print(f"  ❌ {view:35s} | 错误: {e}")

    print("\n" + "=" * 70)
    print("  导入完成!")
    print("=" * 70)


# ============================================================
# 主流程
# ============================================================

def main():
    logger.info("=" * 60)
    logger.info("禅道 Dashboard 数据导入开始")
    logger.info("=" * 60)

    # 1. 初始化数据库
    conn = init_db(DB_PATH)

    # 2. 导入需求
    req_total = 0
    for f in REQ_FILES:
        if not f.exists():
            logger.warning(f"需求文件不存在: {f}")
            continue
        conn.execute("BEGIN TRANSACTION")
        try:
            count = import_requirements(conn, f)
            req_total += count
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.error(f"需求导入失败: {f.name} - {e}")

    # 3. 导入Bug
    bug_total = 0
    for f in BUG_FILES:
        if not f.exists():
            logger.warning(f"Bug文件不存在: {f}")
            continue
        conn.execute("BEGIN TRANSACTION")
        try:
            count = import_bugs(conn, f)
            bug_total += count
            conn.commit()
        except Exception as e:
            conn.rollback()
            logger.error(f"Bug导入失败: {f.name} - {e}")

    # 4. 输出统计报告
    print_report(conn)

    conn.close()
    logger.info(f"数据库已保存: {DB_PATH}")


if __name__ == "__main__":
    main()
