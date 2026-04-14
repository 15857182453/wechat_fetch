#!/usr/bin/env python3
"""
业务对账透视汇总（自动化版）
功能：
1. 从 workspace 目录自动识别并读取最新的 4 个源文件
2. 处理数据并生成 `流水-YYYYMMDD.xlsx`
"""

import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import glob
import os
import re
from pathlib import Path

# ==================== 通用函数 ====================

def from_pivot(pivot_df, 业务类目):
    """从透视表中提取单个业务类目的金额和订单数"""
    row = pivot_df[pivot_df['业绩类目'] == 业务类目]
    if len(row) == 0:
        return 0, 0
    return float(row['实际支付金额_求和'].values[0]), int(row['实际支付金额_计数'].values[0])

def deduct_hospitals(pivot_df, 业务类目, hospital_details):
    """
    从总数据中减去指定医院的数据
    hospital_details: [(医院名称, 金额, 订单), ...]
    """
    total_sum, total_count = from_pivot(pivot_df, 业务类目)
    subtract_sum = sum(h[1] for h in hospital_details)
    subtract_count = sum(h[2] for h in hospital_details)
    return total_sum - subtract_sum, total_count - subtract_count

def add_multiple(items):
    """多个项目金额订单求和（兼容二元组和三元组）"""
    if not items:
        return 0, 0
    # 检测是否为三元组（带名称）
    if len(items[0]) == 3:
        total_amount = sum(float(i[1]) for i in items)
        total_count = sum(int(i[2]) for i in items)
    else:
        total_amount = sum(float(i[0]) for i in items)
        total_count = sum(int(i[1]) for i in items)
    return total_amount, total_count

# ==================== 文件自动识别 ====================

workspace = Path('/home/openclaw/.openclaw/workspace')

def find_latest_file(pattern):
    """根据模式查找最新文件"""
    files = sorted(glob.glob(str(workspace / pattern)), key=os.path.getmtime, reverse=True)
    return files[0] if files else None

# 自动识别 4 个源文件
stats_file = find_latest_file('业务对账统计明细*.xlsx')
if not stats_file:
    raise FileNotFoundError('未找到业务对账统计明细文件（如：业务对账统计明细_20260331.xlsx）')

order_file = find_latest_file('订单明细表*.xlsx')
if not order_file:
    raise FileNotFoundError('未找到订单明细表文件（如：订单明细表-20260331144213.xlsx）')

# 读取日期（从文件名提取）
date_match = re.search(r'(\d{8})', os.path.basename(stats_file))
today_str = date_match.group(1) if date_match else datetime.now().strftime('%Y%m%d')

print(f'\n📊 自动识别文件：')
print(f'  业务对账统计明细: {stats_file}')
print(f'  订单明细表: {order_file}')
print(f'  日期: {today_str}')

# 检查模板文件
template_file = workspace / '流水test.xlsx'
if not template_file.exists():
    raise FileNotFoundError(f'未找到模板文件: {template_file}')

# ==================== 主程序 ====================

# 读取数据
df = pd.read_excel(stats_file)
print('=== 数据透视表逻辑处理 ===\n')

# 筛选逻辑
df收费 = df[df['收退标识'] == '收费'].copy()
exclude_hospitals = ['黑龙江中医药大学附属第一医院', '上海安达医院']
df_filtered = df收费[~df收费['机构名称'].isin(exclude_hospitals)].copy()
df_filtered = df_filtered[~df_filtered['业绩类目'].str.contains('医院咨询|医院复诊|医院护理', na=False)].copy()

# 创建透视表
pivot_table = pd.pivot_table(
    df_filtered, index='业绩类目', values='实际支付金额',
    aggfunc=['sum', 'count'], fill_value=0, margins=True, margins_name='总计'
)
pivot_table.columns = ['实际支付金额_求和', '实际支付金额_计数']
pivot_table = pivot_table.reset_index()

# 保存透视表
output_file = f'/home/openclaw/.openclaw/workspace/业务对账透视汇总_{today_str}.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    pivot_table.to_excel(writer, sheet_name='透视汇总-业绩类目', index=False)

# ==================== 各业务模块 ====================

# 1. 咨询/复诊/护理
print('\n' + '='*80)
print('📊 咨询/复诊/护理（医疗类相关业务）')
print('='*80)
consult_file = '/mnt/c/Users/44238/Desktop/业务对账数据/咨询列表-20260331111624.xlsx'
try:
    df_consult = pd.read_excel(consult_file, sheet_name='咨询')
    df_consult = df_consult[df_consult['支付状态'].isin(['已支付', '退款成功'])]
    consult_amount = df_consult['咨询费用'].sum()
    consult_count = df_consult['咨询费用'].count()
    
    自营_sum, 自营_count = from_pivot(pivot_table, '自营咨询/复诊/护理（医疗类相关业务）')
    total_amount = 自营_sum + consult_amount
    total_count = 自营_count + consult_count
    print(f'自营: {自营_sum:.2f} 元, {自营_count} 笔')
    print(f'外部咨询: {consult_amount:.2f} 元, {consult_count} 笔')
    print(f'📊 合计: {total_amount:.2f} 元, {total_count} 笔')
except Exception as e:
    print(f'⚠️ 错误: {e}')

# 2. 处方服务
print('\n' + '='*80)
print('📊 处方服务（运费 + 便捷购药 + 电子处方）')
print('='*80)
try:
    df_order = pd.read_excel(order_file)
    运费 = (df_order['运费'].sum(), df_order['运费'].count())
    方便购药 = from_pivot(pivot_table, '处方服务-便捷购药')
    电子处方 = from_pivot(pivot_table, '处方服务-电子处方')
    
    处方总金额, 处方总订单 = add_multiple([运费, 方便购药, 电子处方])
    print(f'运费: {运费[0]:.2f} 元, {运费[1]} 笔')
    print(f'便捷购药: {方便购药[0]:.2f} 元, {方便购药[1]} 笔')
    print(f'电子处方: {电子处方[0]:.2f} 元, {电子处方[1]} 笔')
    print(f'📊 合计: {处方总金额:.2f} 元, {处方总订单} 笔')
except Exception as e:
    print(f'⚠️ 错误: {e}')

# 3. 第三方服务（减去三家医院）
print('\n' + '='*80)
print('📊 第三方服务（总数据 - 三家医院）')
print('='*80)
第三方 = from_pivot(pivot_table, '第三方服务')

# 动态获取三家医院的数据
杭州_third = df_filtered[(df_filtered['机构名称'] == '杭州市第七人民医院') & (df_filtered['业绩子类目'] == '第三方其他（病历复印等）')]
杭州_third_sum = float(杭州_third['实际支付金额'].sum()) if len(杭州_third) > 0 else 0.0
杭州_third_count = int(杭州_third['实际支付金额'].count()) if len(杭州_third) > 0 else 0

海宁_four = df_filtered[(df_filtered['机构名称'] == '海宁四院') & (df_filtered['业绩类目'] == '第三方服务')]
海宁_four_sum = float(海宁_four['实际支付金额'].sum()) if len(海宁_four) > 0 else 0.0
海宁_four_count = int(海宁_four['实际支付金额'].count()) if len(海宁_four) > 0 else 0

绍兴_three = df_filtered[(df_filtered['机构名称'] == '绍兴市第七人民医院') & (df_filtered['业绩类目'] == '第三方服务')]
绍兴_three_sum = float(绍兴_three['实际支付金额'].sum()) if len(绍兴_three) > 0 else 0.0
绍兴_three_count = int(绍兴_three['实际支付金额'].count()) if len(绍兴_three) > 0 else 0

三家合计 = (杭州_third_sum + 海宁_four_sum + 绍兴_three_sum, 杭州_third_count + 海宁_four_count + 绍兴_three_count)

第三方服务 = (第三方[0] - 三家合计[0], 第三方[1] - 三家合计[1])
print(f'总数据: {第三方[0]:.2f} 元, {第三方[1]} 笔')
print(f'减: 杭州市第七人民医院 - 业绩子类目 - 第三方其他（病历复印等）: {杭州_third_sum:.2f} 元, {杭州_third_count} 笔')
print(f'减: 海宁四院 - 第三方服务: {海宁_four_sum:.2f} 元, {海宁_four_count} 笔')
print(f'减: 绍兴七院 - 第三方服务: {绍兴_three_sum:.2f} 元, {绍兴_three_count} 笔')
print(f'📊 FINAL: {第三方服务[0]:.2f} 元, {第三方服务[1]} 笔')

# 4. 自营健管（减去绍兴七院）
print('\n' + '='*80)
print('📊 自营健管（总数据 - 绍兴七院）')
print('='*80)
健管 = from_pivot(pivot_table, '自营健管')

# 从原始数据动态计算绍兴七院的自营健管数据
绍兴_data = df_filtered[(df_filtered['机构名称'] == '绍兴市第七人民医院') & (df_filtered['业绩类目'] == '自营健管')]
绍兴_sum = float(绍兴_data['实际支付金额'].sum()) if len(绍兴_data) > 0 else 0.0
绍兴_count = int(绍兴_data['实际支付金额'].count()) if len(绍兴_data) > 0 else 0

健管final = (健管[0] - 绍兴_sum, 健管[1] - 绍兴_count)
print(f'总数据: {健管[0]:.2f} 元, {健管[1]} 笔')
print(f'减: 绍兴七院 {绍兴_sum:.2f} 元, {绍兴_count} 笔')
print(f'📊 FINAL: {健管final[0]:.2f} 元, {健管final[1]} 笔')

# 5. 心理咨询（4项动态相加）
print('\n' + '='*80)
print('📊 心理咨询（4项动态相加）')
print('='*80)

# 杭州市第七人民医院 - 第三方其他（病历复印等）
杭州third = df_filtered[(df_filtered['机构名称'] == '杭州市第七人民医院') & (df_filtered['业绩子类目'] == '第三方其他（病历复印等）')]
杭州_third_sum = float(杭州third['实际支付金额'].sum()) if len(杭州third) > 0 else 0.0
杭州_third_count = int(杭州third['实际支付金额'].count()) if len(杭州third) > 0 else 0

# 海宁四院 - 第三方服务
海宁_four = df_filtered[(df_filtered['机构名称'] == '海宁四院') & (df_filtered['业绩类目'] == '第三方服务')]
海宁_four_sum = float(海宁_four['实际支付金额'].sum()) if len(海宁_four) > 0 else 0.0
海宁_four_count = int(海宁_four['实际支付金额'].count()) if len(海宁_four) > 0 else 0

# 绍兴七院 - 第三方服务
绍兴_three = df_filtered[(df_filtered['机构名称'] == '绍兴市第七人民医院') & (df_filtered['业绩类目'] == '第三方服务')]
绍兴_three_sum = float(绍兴_three['实际支付金额'].sum()) if len(绍兴_three) > 0 else 0.0
绍兴_three_count = int(绍兴_three['实际支付金额'].count()) if len(绍兴_three) > 0 else 0

# 绍兴七院 - 自营健管（同上）
绍兴_self_sum = 绍兴_sum
绍兴_self_count = 绍兴_count

心理金额 = 杭州_third_sum + 海宁_four_sum + 绍兴_three_sum + 绍兴_self_sum
心理订单 = 杭州_third_count + 海宁_four_count + 绍兴_three_count + 绍兴_self_count

print(f'杭州市第七人民医院 - 业绩子类目 - 第三方其他（病历复印等）: {杭州_third_sum:.2f} 元, {杭州_third_count} 笔')
print(f'海宁四院 - 第三方服务: {海宁_four_sum:.2f} 元, {海宁_four_count} 笔')
print(f'绍兴七院 - 第三方服务: {绍兴_three_sum:.2f} 元, {绍兴_three_count} 笔')
print(f'绍兴七院 - 自营健管: {绍兴_self_sum:.2f} 元, {绍兴_self_count} 笔')
print(f'📊 合计: {心理金额:.2f} 元, {心理订单} 笔')

# ==================== 最后打印三个医院的总结 ====================
print('\n' + '='*80)
print('📊 三个医院总结（最终效果）')
print('='*80)
print(f' 杭州七院:{杭州_third_sum:.2f}元,{杭州_third_count}笔')
print(f' 海宁:{海宁_four_sum:.2f}元,{海宁_four_count}笔')
print(f' 绍兴:{绍兴_three_sum:.2f}元,{绍兴_three_count}笔')

# 6. 填入流水-YYYYMMDD.xlsx
print('\n' + '='*80)
print(f'📊 填入流水-{today_str}.xlsx')
print('='*80)
test_file = f'/home/openclaw/.openclaw/workspace/流水-{today_str}.xlsx'
try:
    # 先复制流水test.xlsx作为模板
    template_file = '/home/openclaw/.openclaw/workspace/流水test.xlsx'
    import shutil
    shutil.copy(template_file, test_file)
    
    wb = load_workbook(test_file)
    ws = wb.active
    
    # 列映射
    col_map = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if '会员服务' in str(val): col_map['会员服务'] = col
        elif '支付类' in str(val): col_map['支付类'] = col
        elif '自营体检' in str(val): col_map['自营体检'] = col
        elif '自营健管' in str(val): col_map['自营健管'] = col
        elif '第三方服务' in str(val): col_map['第三方服务'] = col
        elif '心理咨询' in str(val): col_map['心理咨询'] = col
        elif '便捷购药' in str(val): col_map['处方服务'] = col
    
    # 填入数据
    # 动态从透视表读取各业务数据
    体检_sum, 体检_count = from_pivot(pivot_table, '自营体检')
    会员_sum, 会员_count = from_pivot(pivot_table, '会员服务')
    支付_sum, 支付_count = from_pivot(pivot_table, '支付类')

    # 三个医院的总结数据
    杭州七院_data = (杭州_third_sum, 杭州_third_count)
    海宁_data = (海宁_four_sum, 海宁_four_count)
    绍兴_data = (绍兴_three_sum, 绍兴_three_count)

    data_map = {
        '处方服务': (处方总金额, 处方总订单),
        '自营体检': (体检_sum, 体检_count),
        '自营健管': (健管final[0], 健管final[1]),
        '第三方服务': (第三方服务[0], 第三方服务[1]),
        '心理咨询': (心理金额, 心理订单),
        '支付类': (支付_sum, 支付_count),
        '会员服务': (会员_sum, 会员_count),
        '杭州七院': 杭州七院_data,
        '海宁': 海宁_data,
        '绍兴': 绍兴_data
    }
    
    for 业务, (金额, 订单) in data_map.items():
        if 业务 in col_map:
            ws.cell(row=5, column=col_map[业务], value=金额)
            ws.cell(row=5, column=col_map[业务]+1, value=订单)
            print(f'✅ {业务}: {金额:.2f} 元, {订单} 笔')
    
    wb.save(test_file); wb.close()
    print(f'\n✅ 已保存到: {test_file}')

except FileNotFoundError:
    print(f'⚠️ 文件未找到: {test_file} (需要先有流水test.xlsx)')
except Exception as e:
    print(f'⚠️ 错误: {e}')
    import traceback; traceback.print_exc()
