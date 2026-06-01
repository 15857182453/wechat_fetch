#!/usr/bin/env python3
"""
金佰川数据导入脚本 — 标准化流程 v2
流程: 扫描 → 识别 → 预览 → 确认 → 导入(去重) → 验证 → 刷新
"""
import sys, os, re
import psycopg2
import openpyxl
from datetime import datetime
from collections import defaultdict

DB = "host=localhost dbname=jinbaichuan user=openclaw password=jbc2026"
DATA_DIR = "/mnt/e/金佰川数据看板项目"

def conn():
    return psycopg2.connect(DB)

# ══════════════════════════════════════════════════════════════
# 1. 文件扫描 & 识别
# ══════════════════════════════════════════════════════════════
def scan_files():
    """扫描目录，识别文件类型"""
    files = {}
    for f in os.listdir(DATA_DIR):
        path = os.path.join(DATA_DIR, f)
        if not f.endswith('.xlsx'): continue
        size_mb = os.path.getsize(path) / 1024 / 1024

        if '流水号' in f:
            files['detail'] = (f, path, size_mb)
        elif '每日' in f:
            files['daily'] = (f, path, size_mb)
        elif '库存' in f:
            files['inventory'] = (f, path, size_mb)
        elif '上市' in f:
            files['launch'] = (f, path, size_mb)
    return files

def peek_excel(path, max_rows=5):
    """预览 Excel 前几行"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(max_rows+3, ws.max_row), values_only=True)):
        rows.append(list(row))
    total = ws.max_row
    wb.close()
    return rows, total

def identify_detail_date_range(path):
    """识别明细表的日期范围和业务月份"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    dates = set()
    sample_rows = []
    row_count = 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        ts = row[10]  # 单据提交时间
        if ts is not None:
            if isinstance(ts, datetime):
                dates.add(ts.strftime('%Y-%m-%d'))
            elif isinstance(ts, str):
                dates.add(ts[:10])
        row_count += 1
        if len(sample_rows) < 3:
            sample_rows.append(row)
    wb.close()
    return sorted(dates), row_count, sample_rows

# ══════════════════════════════════════════════════════════════
# 2. 预览报告
# ══════════════════════════════════════════════════════════════
def preview(files):
    """生成预览报告，供用户确认"""
    print("\n" + "=" * 70)
    print("📋 数据预览报告")
    print("=" * 70)

    report = {}

    if 'detail' in files:
        fname, path, size = files['detail']
        dates, rows, samples = identify_detail_date_range(path)
        report['detail'] = {
            'file': fname, 'size_mb': size, 'rows': rows,
            'date_min': dates[0] if dates else '?',
            'date_max': dates[-1] if dates else '?',
            'days': len(dates),
            'months': sorted(set(d[:7] for d in dates)) if dates else [],
        }
        r = report['detail']
        print(f"\n📥 交易明细: {fname}")
        print(f"   大小: {size:.1f}MB | 行数: {rows:,} | 日期: {r['date_min']} ~ {r['date_max']} ({r['days']}天)")
        print(f"   涉及月份: {', '.join(r['months'])}")

        # 检查已有数据
        c = conn()
        cur = c.cursor()
        for m in r['months']:
            cur.execute("SELECT COUNT(*) FROM sales_detail WHERE submit_date >= %s AND submit_date < date(%s)+interval '1 month'", (m+'-01', m+'-01'))
            existing = cur.fetchone()[0]
            print(f"   {m}: 数据库中已有 {existing:,} 条 {'⚠️ 将跳过重复' if existing > 0 else '✅ 新数据'}")
        cur.close(); c.close()

    if 'daily' in files:
        fname, path, size = files['daily']
        rows_info, total = peek_excel(path)
        report['daily'] = {'file': fname, 'size_mb': size, 'rows': total}
        print(f"\n📥 日汇总: {fname} ({size:.1f}MB, {total}行)")

    if 'inventory' in files:
        fname, path, size = files['inventory']
        rows_info, total = peek_excel(path)
        report['inventory'] = {'file': fname, 'size_mb': size, 'rows': total}
        print(f"\n📥 库存: {fname} ({size:.1f}MB, {total}行)")

    if 'launch' in files:
        fname, path, size = files['launch']
        rows_info, total = peek_excel(path)
        report['launch'] = {'file': fname, 'size_mb': size, 'rows': total}
        print(f"\n📥 商品上市: {fname} ({size:.1f}MB, {total}行)")

    return report

# ══════════════════════════════════════════════════════════════
# 3. 导入（带去重）
# ══════════════════════════════════════════════════════════════
def import_detail(path, dry_run=False):
    """导入交易明细 — 按(doc_no, product_name)去重"""
    print(f"\n📥 导入交易明细...")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    sql = """
        INSERT INTO sales_detail
            (store_name, doc_no, salesperson, dept_name, brand_name,
             product_name, mnemonic, quantity, settle_amount, gross_profit,
             submit_time, submit_date, hour, is_return)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT DO NOTHING
    """
    # 注意: 需要唯一约束，用(doc_no, product_name, submit_time)做业务键
    # 先建唯一索引（如果不存在）
    c = conn()
    cur = c.cursor()
    cur.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS idx_sd_doc_product_time
        ON sales_detail (doc_no, product_name, submit_time)
    """)
    c.commit()

    batch = []
    total, skipped, imported = 0, 0, 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        store, doc, person, dept, brand, product, mn, qty, amt, profit, ts = row
        if ts is None: continue

        amt_v = float(amt or 0)
        profit_v = float(profit or 0)
        ts_dt = ts if isinstance(ts, datetime) else datetime.strptime(str(ts)[:19], '%Y-%m-%d %H:%M:%S')

        batch.append((
            store, doc, person, dept, brand, product, mn,
            int(qty or 0), amt_v, profit_v,
            ts_dt, ts_dt.date(), ts_dt.hour,
            amt_v < 0
        ))
        total += 1

        if len(batch) >= 5000:
            if not dry_run:
                cur.executemany(sql, batch)
                c.commit()
            batch = []

    if batch and not dry_run:
        cur.executemany(sql, batch)
        c.commit()

    cur.execute("SELECT COUNT(*) FROM sales_detail")
    imported = cur.fetchone()[0]
    cur.close(); c.close(); wb.close()

    print(f"   总数: {total:,} | 数据库现有: {imported:,}")
    return total, imported

def import_daily(path, dry_run=False):
    """导入日汇总 — 按(store_name, sale_date) UPSERT"""
    print(f"\n📥 导入每日销售汇总...")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    c = conn()
    cur = c.cursor()
    total = 0

    for row in ws.iter_rows(min_row=8, max_row=61, values_only=True):
        vals = list(row)
        store = vals[0]
        if not store or store in ('微店', '总计'): continue
        for day in range(1, 32):
            amt = vals[day] if day < len(vals) else None
            if amt is None: continue
            sale_date = f"2026-05-{day:02d}"  # TODO: 从文件头解析月份
            if not dry_run:
                cur.execute("""
                    INSERT INTO sales_daily (store_name, sale_date, total_amount)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (store_name, sale_date)
                    DO UPDATE SET total_amount = EXCLUDED.total_amount, created_at = NOW()
                """, (store, sale_date, float(amt)))
            total += 1

    c.commit(); cur.close(); c.close(); wb.close()
    print(f"   ✅ 导入/更新 {total} 条")
    return total

def import_inventory(path, dry_run=False):
    """导入库存 — 按(brand_name, location, snapshot_date)去重"""
    print(f"\n📥 导入库存...")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    c = conn()
    cur = c.cursor()
    total = 0
    snapshot_date = datetime.now().strftime('%Y-%m-%d')

    for row in ws.iter_rows(min_row=3, values_only=True):
        brand, stock, loc = row[0], row[1], row[2]
        if brand is None: continue
        if not dry_run:
            cur.execute("""
                INSERT INTO inventory_snapshot (brand_name, location, stock_qty, snapshot_date)
                VALUES (%s,%s,%s,%s)
                ON CONFLICT (brand_name, location, snapshot_date)
                DO UPDATE SET stock_qty = EXCLUDED.stock_qty
            """, (str(brand), str(loc), int(stock), snapshot_date))
        total += 1

    c.commit(); cur.close(); c.close(); wb.close()
    print(f"   ✅ 导入/更新 {total} 条 (快照日期: {snapshot_date})")
    return total

def import_launch(path, dry_run=False):
    """导入商品上市 — 全量覆盖（小表）"""
    print(f"\n📥 导入商品上市时间...")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    c = conn()
    cur = c.cursor()
    if not dry_run:
        cur.execute("DELETE FROM prod_launch")  # 小表全量刷新
    total = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        dept, brand, year, season, product, mn, launch = row
        if dept is None: continue
        if not dry_run:
            cur.execute("""
                INSERT INTO prod_launch (dept_name, brand_name, launch_year, season, product_name, mnemonic, launch_date)
                VALUES (%s,%s,%s,%s,%s,%s,%s)
            """, (str(dept), str(brand or ''), int(year or 0), str(season or ''),
                  str(product or ''), str(mn or ''), launch.date() if launch else None))
        total += 1

    c.commit(); cur.close(); c.close(); wb.close()
    print(f"   ✅ 导入 {total} 条")
    return total

def refresh_dimensions():
    """刷新维度表"""
    print(f"\n🔄 刷新维度表...")
    c = conn()
    cur = c.cursor()
    dims = [
        ("dim_store", "store_name", "sales_detail", "store_name"),
        ("dim_brand", "brand_name", "sales_detail", "brand_name"),
        ("dim_dept", "dept_name", "sales_detail", "dept_name"),
    ]
    for table, col, source, src_col in dims:
        cur.execute(f"""
            INSERT INTO {table} ({col})
            SELECT DISTINCT {src_col} FROM {source}
            ON CONFLICT DO NOTHING
        """)
        cur.execute(f"SELECT COUNT(*) FROM {table}")
        print(f"   {table}: {cur.fetchone()[0]} 条")
    c.commit(); cur.close(); c.close()

def refresh_materialized_views():
    """刷新物化视图"""
    print(f"\n🔄 刷新物化视图...")
    c = conn()
    cur = c.cursor()
    for mv in ['mv_brand_daily', 'mv_store_daily', 'mv_dept_daily']:
        cur.execute(f"REFRESH MATERIALIZED VIEW {mv}")
        cur.execute(f"SELECT COUNT(*) FROM {mv}")
        print(f"   {mv}: {cur.fetchone()[0]} 行")
    c.commit(); cur.close(); c.close()

def verify():
    """验证导入结果"""
    print(f"\n📊 数据验证:")
    c = conn()
    cur = c.cursor()
    cur.execute("""
        SELECT '交易明细' as tbl, COUNT(*)::text FROM sales_detail
        UNION ALL SELECT '日汇总', COUNT(*)::text FROM sales_daily
        UNION ALL SELECT '库存', COUNT(*)::text FROM inventory_snapshot
        UNION ALL SELECT '上市商品', COUNT(*)::text FROM prod_launch
        UNION ALL SELECT '品牌', COUNT(*)::text FROM dim_brand
        UNION ALL SELECT '门店', COUNT(*)::text FROM dim_store
        UNION ALL SELECT '品类', COUNT(*)::text FROM dim_dept
        UNION ALL SELECT '品牌日聚合', COUNT(*)::text FROM mv_brand_daily
        UNION ALL SELECT '门店日聚合', COUNT(*)::text FROM mv_store_daily
    """)
    for tbl, cnt in cur.fetchall():
        print(f"   {tbl}: {cnt}")

    # 关键 KPI 验证
    cur.execute("""
        SELECT SUM(settle_amount), SUM(gross_profit), COUNT(DISTINCT doc_no)
        FROM sales_detail WHERE NOT is_return AND submit_date >= '2026-05-01'
    """)
    amt, profit, orders = cur.fetchone()
    if amt:
        print(f"\n   月销额: ¥{float(amt)/10000:.1f}万")
        print(f"   毛利率: {float(profit)/float(amt)*100:.1f}%")
        print(f"   订单数: {orders:,}")
    cur.close(); c.close()

# ══════════════════════════════════════════════════════════════
# 主流程
# ══════════════════════════════════════════════════════════════
if __name__ == '__main__':
    args = sys.argv[1:]
    dry_run = '--dry-run' in args
    auto_confirm = '--yes' in args

    print("🔍 扫描数据目录...")
    files = scan_files()

    if not files:
        print(f"❌ 未在 {DATA_DIR} 找到数据文件")
        sys.exit(1)

    report = preview(files)

    if dry_run:
        print("\n⚠️ DRY RUN 模式 — 仅预览，不导入")
        sys.exit(0)

    if not auto_confirm:
        print("\n" + "=" * 70)
        resp = input("确认导入以上数据? [Y/n]: ").strip().lower()
        if resp and resp not in ('y', 'yes', '是'):
            print("已取消")
            sys.exit(0)

    # 执行导入
    if 'detail' in files:
        import_detail(files['detail'][1])
    if 'daily' in files:
        import_daily(files['daily'][1])
    if 'inventory' in files:
        import_inventory(files['inventory'][1])
    if 'launch' in files:
        import_launch(files['launch'][1])

    refresh_dimensions()
    refresh_materialized_views()
    verify()

    print("\n✅ 导入完成！")
