#!/usr/bin/env python3
"""
业务对账自动化脚本（简单版）
三列表格：业务分类 | 金额（元） | 订单
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
import glob
import os
import re
import sys

def find_latest_file(dir_path, pattern):
    files = sorted(glob.glob(os.path.join(dir_path, pattern)), key=os.path.getmtime, reverse=True)
    return files[0] if files else None

def from_pivot(pivot_df, 业务类目):
    row = pivot_df[pivot_df['业绩类目'] == 业务类目]
    if len(row) == 0:
        return 0.0, 0
    return float(row['金额'].values[0]), int(row['订单'].values[0])

print('='*80)
print('📊 业务对账自动化脚本（简单版）')
print('='*80)

current_dir = os.getcwd()

stats_file = find_latest_file(current_dir, '业务对账统计明细*.xlsx')
order_file = find_latest_file(current_dir, '订单明细表*.xlsx')
consult_files = glob.glob(os.path.join(current_dir, '*咨询*.xlsx'))

if not stats_file:
    print('❌ 未找到业务对账统计明细文件')
    sys.exit(1)
if not order_file:
    print('❌ 未找到订单明细表文件')
    sys.exit(1)

date_match = re.search(r'(\d{8})', os.path.basename(stats_file))
today_str = date_match.group(1) if date_match else datetime.now().strftime('%Y%m%d')

print(f'\n📊 文件: {os.path.basename(stats_file)}')
print(f'📅 日期: {today_str}')

df = pd.read_excel(stats_file)

df收费 = df[df['收退标识'] == '收费'].copy()
df_filtered = df收费[~df收费['机构名称'].isin(['黑龙江中医药大学附属第一医院', '上海安达医院'])].copy()
df_filtered = df_filtered[~df_filtered['业绩类目'].str.contains('医院咨询|医院复诊|医院护理', na=False)].copy()

pivot = pd.pivot_table(df_filtered, index='业绩类目', values='实际支付金额', aggfunc=['sum', 'count'])
pivot.columns = ['金额', '订单']
pivot = pivot.reset_index()

# 动态获取医院数据
杭州_third = df_filtered[(df_filtered['机构名称'] == '杭州市第七人民医院') & (df_filtered['业绩子类目'] == '第三方其他（病历复印等）')]
杭州_third_sum = float(杭州_third['实际支付金额'].sum()) if len(杭州_third) > 0 else 0.0
杭州_third_count = int(杭州_third['实际支付金额'].count()) if len(杭州_third) > 0 else 0

海宁_four = df_filtered[(df_filtered['机构名称'] == '海宁四院') & (df_filtered['业绩类目'] == '第三方服务')]
海宁_four_sum = float(海宁_four['实际支付金额'].sum()) if len(海宁_four) > 0 else 0.0
海宁_four_count = int(海宁_four['实际支付金额'].count()) if len(海宁_four) > 0 else 0

绍兴_three = df_filtered[(df_filtered['机构名称'] == '绍兴市第七人民医院') & (df_filtered['业绩类目'] == '第三方服务')]
绍兴_three_sum = float(绍兴_three['实际支付金额'].sum()) if len(绍兴_three) > 0 else 0.0
绍兴_three_count = int(绍兴_three['实际支付金额'].count()) if len(绍兴_three) > 0 else 0

绍兴_self = df_filtered[(df_filtered['机构名称'] == '绍兴市第七人民医院') & (df_filtered['业绩类目'] == '自营健管')]
绍兴_self_sum = float(绍兴_self['实际支付金额'].sum()) if len(绍兴_self) > 0 else 0.0
绍兴_self_count = int(绍兴_self['实际支付金额'].count()) if len(绍兴_self) > 0 else 0

# 计算业务
咨询金额, 咨询订单 = 0, 0
if consult_files:
    try:
        df_consult = pd.read_excel(consult_files[0], sheet_name='咨询')
        df_consult = df_consult[df_consult['支付状态'].isin(['已支付', '退款成功'])]
        咨询金额 = df_consult['咨询费用'].sum()
        咨询订单 = df_consult['咨询费用'].count()
    except:
        pass
自营_sum, 自营_count = from_pivot(pivot, '自营咨询/复诊/护理（医疗类相关业务）')
咨询合计 = (自营_sum + 咨询金额, 自营_count + 咨询订单)

df_order = pd.read_excel(order_file)
运费 = (df_order['运费'].sum(), df_order['运费'].count())

# 宁夏医科大学总医院的运费数据（运费列非空的行数）
宁夏运费 = df_order[df_order['开方机构'] == '宁夏医科大学总医院']['运费']
宁夏运费_count = 宁夏运费.count()

方便购药 = from_pivot(pivot, '处方服务-便捷购药')
电子处方 = from_pivot(pivot, '处方服务-电子处方')
处方服务 = (运费[0] + 方便购药[0] + 电子处方[0], 运费[1] + 方便购药[1] + 电子处方[1])

第三方 = from_pivot(pivot, '第三方服务')
第三合计 = (杭州_third_sum + 海宁_four_sum + 绍兴_three_sum, 杭州_third_count + 海宁_four_count + 绍兴_three_count)
第三方服务 = (第三方[0] - 第三合计[0], 第三方[1] - 第三合计[1])

健管 = from_pivot(pivot, '自营健管')
自营健管 = (健管[0] - 绍兴_self_sum, 健管[1] - 绍兴_self_count)

心理咨询 = (杭州_third_sum + 海宁_four_sum + 绍兴_three_sum + 绍兴_self_sum,
            杭州_third_count + 海宁_four_count + 绍兴_three_count + 绍兴_self_count)

自营体检 = from_pivot(pivot, '自营体检')
会员服务 = from_pivot(pivot, '会员服务')
支付类 = from_pivot(pivot, '支付类')

# 生成文件（简单三列表格）
output_file = os.path.join(current_dir, f'流水-{today_str}.xlsx')
wb = Workbook()
ws = wb.active

# 标题
ws['A1'] = '业务分类'; ws['B1'] = '金额（元）'; ws['C1'] = '订单'

# 数据
ws['A2'] = '咨询/复诊/护理'; ws['B2'] = 咨询合计[0]; ws['C2'] = 咨询合计[1]
ws['A3'] = '处方服务'; ws['B3'] = 处方服务[0]; ws['C3'] = 处方服务[1]
ws['A4'] = '自营体检'; ws['B4'] = 自营体检[0]; ws['C4'] = 自营体检[1]
ws['A5'] = '自营健管'; ws['B5'] = 自营健管[0]; ws['C5'] = 自营健管[1]
ws['A6'] = '第三方服务'; ws['B6'] = 第三方服务[0]; ws['C6'] = 第三方服务[1]
ws['A7'] = '心理咨询'; ws['B7'] = 心理咨询[0]; ws['C7'] = 心理咨询[1]
ws['A8'] = '支付类'; ws['B8'] = 支付类[0]; ws['C8'] = 支付类[1]
ws['A9'] = '会员服务'; ws['B9'] = 会员服务[0]; ws['C9'] = 会员服务[1]

wb.save(output_file)
wb.close()

# ==================== 查询指定医院的处方服务数据 ====================
医院列表 = [
    ('浙江省中医院（湖滨院区）', '处方服务-电子处方'),
    ('杭州师范大学附属医院', '处方服务-便捷购药'),
    ('青岛市中医医院（市海慈医院）', '处方服务-便捷购药'),
    ('齐鲁德医', '处方服务-电子处方'),
    ('山东大学齐鲁第二医院(中心院区)', '处方服务-电子处方'),
    ('安徽省立医院', '处方服务-电子处方'),
    ('青岛中心医院', '处方服务-电子处方')
]

处方服务医院数据 = []
for 医院, 类目 in 医院列表:
    # 使用 strip 后的名称进行比较，处理机构名称前后有空格的情况
    医院_data = df_filtered[
        (df_filtered['机构名称'].str.strip() == 医院) &
        (df_filtered['收退标识'] == '收费') &
        (df_filtered['业绩类目'] == 类目)
    ]
    金额 = float(医院_data['实际支付金额'].sum()) if len(医院_data) > 0 else 0.0
    订单数 = int(医院_data['实际支付金额'].count()) if len(医院_data) > 0 else 0
    处方服务医院数据.append((医院, 金额, 订单数, 类目))
    print(f' {医院}:{金额:.2f}元,{订单数}笔({类目})')

# 宁夏医科大学总医院运费数据
print(f' 宁夏医科大学总医院运费订单: {宁夏运费_count}笔')

# ==================== 打印三个医院的总结 ====================
# 宁夏医科大学总医院运费数据
print(f' 宁夏医科大学总医院运费订单: {宁夏运费_count}笔')

print(f'\n📊 三个医院总结:')
print(f' 杭州七院:{杭州_third_sum:.2f}元,{杭州_third_count}笔')
print(f' 海宁:{海宁_four_sum:.2f}元,{海宁_four_count}笔')
print(f' 绍兴:{绍兴_self_sum:.2f}元,{绍兴_self_count}笔')

print(f'\n✅ 已生成: {output_file}')
print(f'\n📊 数据:')
print(f'  咨询/复诊/护理: {咨询合计[0]:.2f}元, {咨询合计[1]}笔')
print(f'  处方服务: {处方服务[0]:.2f}元, {处方服务[1]}笔')
print(f'  自营体检: {自营体检[0]:.2f}元, {自营体检[1]}笔')
print(f'  自营健管: {自营健管[0]:.2f}元, {自营健管[1]}笔')
print(f'  第三方服务: {第三方服务[0]:.2f}元, {第三方服务[1]}笔')
print(f'  心理咨询: {心理咨询[0]:.2f}元, {心理咨询[1]}笔')
print(f'  支付类: {支付类[0]:.2f}元, {支付类[1]}笔')
print(f'  会员服务: {会员服务[0]:.2f}元, {会员服务[1]}笔')

print(f'  宁夏医科大学总医院运费订单: {宁夏运费_count}笔')

# ==================== 在Excel中填入三个医院的数据（追加到右边） ====================
from openpyxl import load_workbook

wb = load_workbook(output_file)
ws = wb.active

# 在E列开始填入医院数据
ws['E1'] = '医院'; ws['F1'] = '金额（元）'; ws['G1'] = '订单'
ws['E2'] = '杭州七院'; ws['F2'] = 杭州_third_sum; ws['G2'] = 杭州_third_count
ws['E3'] = '海宁'; ws['F3'] = 海宁_four_sum; ws['G3'] = 海宁_four_count
ws['E4'] = '绍兴'; ws['F4'] = 绍兴_three_sum; ws['G4'] = 绍兴_three_count

# 在K列填入宁夏医科大学总医院运费数据
ws['K1'] = '医院'; ws['L1'] = '运费订单数'
ws['K2'] = '宁夏医科大学总医院'; ws['L2'] = 宁夏运费_count

# 在H列开始填入处方服务医院数据
ws['H1'] = '医院'; ws['I1'] = '金额（元）'; ws['J1'] = '订单'
for i, (医院, 金额, 订单数, 类目) in enumerate(处方服务医院数据, start=2):
    ws[f'H{i}'] = f'{医院.strip()}({类目})'
    ws[f'I{i}'] = 金额
    ws[f'J{i}'] = 订单数

wb.save(output_file)
wb.close()

# ==================== 复制到桌面老位置 ====================
# Windows桌面挂载路径
desktop_dir = '/mnt/c/Users/44238/Desktop/对账测试3-29数据'
if os.path.exists(desktop_dir):
    import shutil
    dest_file = os.path.join(desktop_dir, os.path.basename(output_file))
    shutil.copy2(output_file, dest_file)
    print(f'\n✅ 已复制到桌面: {dest_file}')
