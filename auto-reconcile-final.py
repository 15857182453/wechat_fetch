#!/usr/bin/env python3
"""
业务对账自动化脚本（最终版 - 横向格式）
每业务占2列，8个业务共16列
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
import glob
import os
import re
import sys

def find_latest_file(dir_path, pattern):
    files = sorted(glob.glob(os.path.join(dir_path, pattern)), key=os.path.getmtime, reverse=True)
    return files[0] if files else None

def from_pivot(pivot_df, 业务类目):
    row = pivot_df[pivot_df['业绩类目'] == 业务类目]
    if len(row) == 0:
        return 0.0, 0
    return float(row['金额'].values[0]), int(row['订单'].values[0])

print('='*80)
print('📊 业务对账自动化脚本')
print('='*80)

current_dir = os.getcwd()

stats_file = find_latest_file(current_dir, '业务对账统计明细*.xlsx')
order_file = find_latest_file(current_dir, '订单明细表*.xlsx')
consult_files = glob.glob(os.path.join(current_dir, '*咨询*.xlsx'))

if not stats_file:
    print('❌ 未找到业务对账统计明细文件')
    sys.exit(1)
if not order_file:
    print('❌ 未找到订单明细表文件')
    sys.exit(1)

date_match = re.search(r'(\d{8})', os.path.basename(stats_file))
today_str = date_match.group(1) if date_match else datetime.now().strftime('%Y%m%d')

print(f'\n📊 文件: {os.path.basename(stats_file)}')
print(f'📅 日期: {today_str}')

df = pd.read_excel(stats_file)

df收费 = df[df['收退标识'] == '收费'].copy()
df_filtered = df收费[~df收费['机构名称'].isin(['黑龙江中医药大学附属第一医院', '上海安达医院'])].copy()
df_filtered = df_filtered[~df_filtered['业绩类目'].str.contains('医院咨询|医院复诊|医院护理', na=False)].copy()

pivot = pd.pivot_table(df_filtered, index='业绩类目', values='实际支付金额', aggfunc=['sum', 'count'])
pivot.columns = ['金额', '订单']
pivot = pivot.reset_index()

# 动态获取医院数据
杭州_third = df_filtered[(df_filtered['机构名称'] == '杭州市第七人民医院') & (df_filtered['业绩子类目'] == '第三方其他（病历复印等）')]
杭州_third_sum = float(杭州_third['实际支付金额'].sum()) if len(杭州_third) > 0 else 0.0
杭州_third_count = int(杭州_third['实际支付金额'].count()) if len(杭州_third) > 0 else 0

海宁_four = df_filtered[(df_filtered['机构名称'] == '海宁四院') & (df_filtered['业绩类目'] == '第三方服务')]
海宁_four_sum = float(海宁_four['实际支付金额'].sum()) if len(海宁_four) > 0 else 0.0
海宁_four_count = int(海宁_four['实际支付金额'].count()) if len(海宁_four) > 0 else 0

绍兴_three = df_filtered[(df_filtered['机构名称'] == '绍兴市第七人民医院') & (df_filtered['业绩类目'] == '第三方服务')]
绍兴_three_sum = float(绍兴_three['实际支付金额'].sum()) if len(绍兴_three) > 0 else 0.0
绍兴_three_count = int(绍兴_three['实际支付金额'].count()) if len(绍兴_three) > 0 else 0

绍兴_self = df_filtered[(df_filtered['机构名称'] == '绍兴市第七人民医院') & (df_filtered['业绩类目'] == '自营健管')]
绍兴_self_sum = float(绍兴_self['实际支付金额'].sum()) if len(绍兴_self) > 0 else 0.0
绍兴_self_count = int(绍兴_self['实际支付金额'].count()) if len(绍兴_self) > 0 else 0

# 计算业务
咨询金额, 咨询订单 = 0, 0
if consult_files:
    try:
        df_consult = pd.read_excel(consult_files[0], sheet_name='咨询')
        df_consult = df_consult[df_consult['支付状态'].isin(['已支付', '退款成功'])]
        咨询金额 = df_consult['咨询费用'].sum()
        咨询订单 = df_consult['咨询费用'].count()
    except:
        pass
自营_sum, 自营_count = from_pivot(pivot, '自营咨询/复诊/护理（医疗类相关业务）')
咨询合计 = (自营_sum + 咨询金额, 自营_count + 咨询订单)

df_order = pd.read_excel(order_file)
运费 = (df_order['运费'].sum(), df_order['运费'].count())
方便购药 = from_pivot(pivot, '处方服务-便捷购药')
电子处方 = from_pivot(pivot, '处方服务-电子处方')
处方服务 = (运费[0] + 方便购药[0] + 电子处方[0], 运费[1] + 方便购药[1] + 电子处方[1])

第三方 = from_pivot(pivot, '第三方服务')
第三合计 = (杭州_third_sum + 海宁_four_sum + 绍兴_three_sum, 杭州_third_count + 海宁_four_count + 绍兴_three_count)
第三方服务 = (第三方[0] - 第三合计[0], 第三合计[1] - 第三合计[1])

健管 = from_pivot(pivot, '自营健管')
自营健管 = (健管[0] - 绍兴_self_sum, 健管[1] - 绍兴_self_count)

心理咨询 = (杭州_third_sum + 海宁_four_sum + 绍兴_three_sum + 绍兴_self_sum,
            杭州_third_count + 海宁_four_count + 绍兴_three_count + 绍兴_self_count)

自营体检 = from_pivot(pivot, '自营体检')
会员服务 = from_pivot(pivot, '会员服务')
支付类 = from_pivot(pivot, '支付类')

# 生成文件（横向格式：每业务占2列）
output_file = os.path.join(current_dir, f'流水-{today_str}.xlsx')
wb = Workbook()
ws = wb.active

# 每个业务占2列（金额列 + 订单列）
# 列: A B C D E F G H I J K L M N O P
# 业务: -- 咨询 -- 处方 -- 体检 -- 健管 -- 第三方 -- 心理 -- 支付 -- 会员 --
# 行1: 空 业务名 空 空 业务名 空 空 业务名 空 空 业务名 空 空 业务名 空 空 业务名 空 空
# 行2: 空 空 空 空 空 空 空 空 空 空 空 空 空 空 空 空 空 空 空 空
# 行3: 日期 金额 空 空 金额 空 空 金额 空 空 金额 空 空 金额 空 空 金额 空 空
# 行4: 空 订单 空 空 订单 空 空 订单 空 空 订单 空 空 订单 空 空 订单 空 空

# 行1：业务名称横向排列（每个业务占2列）
ws['B1'] = '咨询/复诊/护理（医疗类相关业务）'
ws['E1'] = '处方服务\n(便捷购药)'; ws['H1'] = '自营体检'; ws['K1'] = '自营健管'
ws['N1'] = '第三方服务'; ws['Q1'] = '心理咨询'; ws['T1'] = '支付类'; ws['W1'] = '会员服务'

# 行2：空行

# 行3：金额横向排列
ws['A3'] = datetime.strptime(today_str, '%Y%m%d')
ws['B3'] = 咨询合计[0]; ws['E3'] = 处方服务[0]; ws['H3'] = 自营体检[0]; ws['K3'] = 自营健管[0]
ws['N3'] = 第三方服务[0]; ws['Q3'] = 心理咨询[0]; ws['T3'] = 支付类[0]; ws['W3'] = 会员服务[0]

# 行4：订单横向排列
ws['B4'] = 咨询合计[1]; ws['E4'] = 处方服务[1]; ws['H4'] = 自营体检[1]; ws['K4'] = 自营健管[1]
ws['N4'] = 第三方服务[1]; ws['Q4'] = 心理咨询[1]; ws['T4'] = 支付类[1]; ws['W4'] = 会员服务[1]

wb.save(output_file)
wb.close()

print(f'\n📊 横向数据:')
print(f'  咨询/复诊/护理: A3={咨询合计[0]:.2f}元, A4={咨询合计[1]}笔')
print(f'  处方服务: E3={处方服务[0]:.2f}元, E4={处方服务[1]}笔')
print(f'  自营体检: H3={自营体检[0]:.2f}元, H4={自营体检[1]}笔')
print(f'  自营健管: K3={自营健管[0]:.2f}元, K4={自营健管[1]}笔')
print(f'  第三方服务: N3={第三方服务[0]:.2f}元, N4={第三方服务[1]}笔')
print(f'  心理咨询: Q3={心理咨询[0]:.2f}元, Q4={心理咨询[1]}笔')
print(f'  支付类: T3={支付类[0]:.2f}元, T4={支付类[1]}笔')
print(f'  会员服务: W3={会员服务[0]:.2f}元, W4={会员服务[1]}笔')
print(f'\n✅ 已生成: {output_file}')
